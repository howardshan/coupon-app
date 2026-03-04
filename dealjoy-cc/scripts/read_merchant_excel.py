#!/usr/bin/env python3
import json, sys, os
try:
    import openpyxl
except ImportError:
    os.system('pip install openpyxl --break-system-packages -q')
    import openpyxl

def read_module(ep, mk):
    wb = openpyxl.load_workbook(ep, data_only=True)
    ws = wb['商家端详细需求']
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rd = dict(zip(headers, row))
        mc = rd.get('模块', '')
        if mc and mk in str(mc):
            rows.append(rd)
    cm = []
    if '注册' in mk or '认证' in mk:
        ws3 = wb['注册证件矩阵']
        ch = [c.value for c in ws3[1]]
        for row in ws3.iter_rows(min_row=2, values_only=True):
            cm.append(dict(zip(ch, row)))
    wb.close()
    if not rows:
        print(f'Not found: {mk}', file=sys.stderr)
        sys.exit(1)
    out = [f'## Merchant Module: {mk}', f'Total {len(rows)} requirements']
    for i, r in enumerate(rows):
        ne = {k: v for k, v in r.items() if v is not None}
        out.append(f'[{i+1}] ' + json.dumps(ne, ensure_ascii=False))
    if cm:
        out.append('')
        out.append('### Certificate Matrix')
        for c in cm:
            ne = {k: v for k, v in c.items() if v is not None}
            out.append(json.dumps(ne, ensure_ascii=False))
    return chr(10).join(out)

def find_excel():
    for p in ['requirements/DealJoy_商家端需求清单.xlsx']:
        if os.path.exists(p): return p
    return None

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print('Usage: python3 scripts/read_merchant_excel.py <module>')
        sys.exit(1)
    mk = sys.argv[1]
    ep = sys.argv[2] if len(sys.argv) > 2 else find_excel()
    if not ep:
        print('Excel not found', file=sys.stderr)
        sys.exit(1)
    print(read_module(ep, mk))
