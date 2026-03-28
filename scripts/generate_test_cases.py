#!/usr/bin/env python3
"""生成 DealJoy 用户端完整测试用例 Excel 表"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ===== 样式定义 =====
header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
module_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
module_font = Font(name='Arial', bold=True, size=11, color='1F3864')
normal_font = Font(name='Arial', size=10)
pass_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
fail_fill = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='B4C6E7'),
    right=Side(style='thin', color='B4C6E7'),
    top=Side(style='thin', color='B4C6E7'),
    bottom=Side(style='thin', color='B4C6E7')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# ===== 列定义 =====
columns = [
    ('编号', 8),
    ('模块', 14),
    ('子模块', 14),
    ('测试场景', 30),
    ('前置条件', 25),
    ('测试步骤', 40),
    ('预期结果', 35),
    ('测试方式', 12),
    ('优先级', 8),
    ('测试结果', 10),
    ('备注', 20),
]

# ===== 测试用例数据 =====
test_cases = [
    # ========== AUTH 模块 ==========
    # 欢迎页
    ('Auth', '欢迎页', '首次打开App显示欢迎页', '未登录状态',
     '1. 清除登录态\n2. 重新打开App',
     '显示 WelcomeScreen，含 "Get Started" 和 "Create Account" 按钮', '手动', 'P0'),
    ('Auth', '欢迎页', '点击 Get Started 跳转登录', '在欢迎页',
     '点击 "Get Started" 按钮',
     '跳转到 LoginScreen', '手动', 'P0'),
    ('Auth', '欢迎页', '点击 Create Account 跳转注册', '在欢迎页',
     '点击 "Create Account" 按钮',
     '跳转到 RegisterScreen', '手动', 'P0'),

    # 登录
    ('Auth', '邮箱登录', '正确邮箱密码登录', '已注册且邮箱已验证',
     '1. 输入正确邮箱\n2. 输入正确密码\n3. 点击 Sign In',
     '登录成功，跳转到 /home', '手动', 'P0'),
    ('Auth', '邮箱登录', '错误密码登录', '已注册用户',
     '1. 输入正确邮箱\n2. 输入错误密码\n3. 点击 Sign In',
     '密码框下方显示红色错误提示 "Invalid email or password"', '手动', 'P0'),
    ('Auth', '邮箱登录', '未注册邮箱登录', '无',
     '1. 输入未注册的邮箱\n2. 输入任意密码\n3. 点击 Sign In',
     '显示红色错误提示 "Invalid email or password"（不泄露邮箱是否存在）', '手动', 'P1'),
    ('Auth', '邮箱登录', '邮箱未验证时登录', '已注册但邮箱未验证',
     '1. 输入未验证邮箱\n2. 输入正确密码\n3. 点击 Sign In',
     '弹出红色提示 "Please verify your email..."，自动登出', '手动', 'P0'),
    ('Auth', '邮箱登录', '空邮箱提交', '在登录页',
     '不输入邮箱，点击 Sign In',
     '邮箱字段显示验证错误提示', '手动', 'P1'),
    ('Auth', '邮箱登录', '邮箱格式错误', '在登录页',
     '输入 "abc@" 格式不完整的邮箱',
     '邮箱字段显示格式错误提示', '手动', 'P1'),
    ('Auth', '邮箱登录', '密码少于8字符', '在登录页',
     '输入有效邮箱 + 7位密码',
     '密码字段显示 "至少8字符" 验证错误', '手动', 'P1'),
    ('Auth', '邮箱登录', '密码显示/隐藏切换', '在登录页',
     '输入密码后点击眼睛图标',
     '密码在明文和密文间切换', '手动', 'P2'),
    ('Auth', '邮箱登录', 'Remember me 选项', '在登录页',
     '勾选/取消 "Keep me signed in for 30 days"',
     '复选框状态正确切换', '手动', 'P2'),
    ('Auth', '邮箱登录', 'Browse as Guest', '在登录页',
     '点击 "Browse as Guest"',
     '跳转到 /home 首页，以游客模式浏览', '手动', 'P1'),
    ('Auth', '邮箱登录', '登录中按钮禁用', '在登录页',
     '点击 Sign In 后立即再次点击',
     '按钮显示加载动画且不可重复点击', '手动', 'P1'),

    # Google 登录
    ('Auth', 'Google 登录', 'Google 账号登录成功', '已安装 Google 服务',
     '点击 "Continue with Google"，选择 Google 账号',
     '授权后自动登录，跳转到 /home', '手动', 'P0'),
    ('Auth', 'Google 登录', '取消 Google 授权', '在 Google 弹窗中',
     '点击 "Continue with Google" 后取消授权',
     '返回登录页，无错误提示', '手动', 'P1'),

    # Apple 登录
    ('Auth', 'Apple 登录', 'Apple 账号登录（iOS）', 'iOS 设备',
     '点击 "Continue with Apple"，完成 Face ID/密码验证',
     '授权后自动登录，跳转到 /home', '手动', 'P0'),
    ('Auth', 'Apple 登录', 'Android 不显示 Apple 登录', 'Android 设备',
     '查看登录页',
     '不显示 "Continue with Apple" 按钮', '手动', 'P1'),

    # 注册
    ('Auth', '注册', '完整流程注册成功', '无',
     '1. 填写 Username（字母数字，2-30位）\n2. 填写 Full Name\n3. 填写有效邮箱\n4. 填写强密码（8+位含大小写和数字）\n5. 确认密码\n6. 勾选 Terms of Service\n7. 点击 Create Account',
     '注册成功，自动登出，跳转到 OTP 验证页', '手动', 'P0'),
    ('Auth', '注册', 'Username 含特殊字符', '在注册页',
     '输入 "user@name"',
     '显示 "Only letters and numbers allowed" 错误', '手动', 'P1'),
    ('Auth', '注册', 'Username 少于2字符', '在注册页',
     '输入 "a"',
     '显示长度不足错误', '手动', 'P1'),
    ('Auth', '注册', '密码强度指示器', '在注册页',
     '依次输入：1234（弱）→ Abcd1234（中）→ Abcd1234!@#（强）',
     '密码强度指示器依次显示 Weak(红) → Medium(橙) → Strong(绿)', '手动', 'P1'),
    ('Auth', '注册', '确认密码不匹配', '在注册页',
     '密码填 "Abc12345"，确认密码填 "Abc12346"',
     '实时显示红色 "Passwords do not match" 提示', '手动', 'P1'),
    ('Auth', '注册', '未勾选 ToS 时按钮禁用', '在注册页',
     '所有字段填好但不勾选 Terms of Service',
     '"Create Account" 按钮灰色不可点击', '手动', 'P1'),
    ('Auth', '注册', '已存在邮箱注册', '在注册页',
     '使用已注册的邮箱注册',
     'SnackBar 显示邮箱已存在的错误信息', '手动', 'P1'),

    # OTP 验证
    ('Auth', 'OTP 验证', '正确验证码验证成功', '注册后跳转到 OTP 页',
     '1. 查看注册邮箱收到的验证码\n2. 输入正确验证码\n3. 点击 Verify',
     '验证成功，绿色 SnackBar 提示，自动跳转 /home', '手动', 'P0'),
    ('Auth', 'OTP 验证', '输入错误验证码', '在 OTP 验证页',
     '输入错误的 6 位数字，点击 Verify',
     '显示红色错误 "Invalid or expired code..."', '手动', 'P0'),
    ('Auth', 'OTP 验证', '重发验证码', '在 OTP 验证页',
     '点击 "Didn\'t receive the code? Resend"',
     '绿色 SnackBar 提示 "Verification code resent!"', '手动', 'P1'),
    ('Auth', 'OTP 验证', '频繁尝试触发限制', '在 OTP 验证页',
     '连续多次输入错误验证码',
     '显示 "Too many attempts. Please wait..." 限制提示', '手动', 'P2'),

    # 忘记密码
    ('Auth', '忘记密码', '发送重置链接', '在登录页',
     '1. 点击 "Forgot password?"\n2. 输入注册邮箱\n3. 点击 "Send Reset Link"',
     '切换到成功视图，显示 60 秒倒计时重发按钮', '手动', 'P0'),
    ('Auth', '忘记密码', '重发链接倒计时', '在发送成功视图',
     '等待倒计时从 60 秒开始递减',
     '倒计时进行中 "Resend Link" 按钮禁用，结束后启用', '手动', 'P1'),
    ('Auth', '忘记密码', '输入未注册邮箱', '在忘记密码页',
     '输入不存在的邮箱，点击发送',
     '仍显示成功视图（隐私保护，不泄露邮箱是否存在）', '手动', 'P1'),

    # 重置密码
    ('Auth', '重置密码', '通过邮件链接重置密码', '收到重置密码邮件',
     '1. 点击邮件中重置链接\n2. App 跳转到 ResetPasswordScreen\n3. 输入新密码（满足强度要求）\n4. 确认密码\n5. 点击 Reset Password',
     '密码更新成功，显示成功视图，3秒后自动跳转登录页', '手动', 'P0'),
    ('Auth', '重置密码', '密码要求实时校验', '在重置密码页',
     '逐步输入密码，观察要求列表',
     '每条要求（8字符/大写/小写/数字）实时变绿', '手动', 'P1'),

    # 登出
    ('Auth', '登出', '正常登出', '已登录状态',
     'Profile 页点击 "Sign Out"',
     '清除 session，跳转到欢迎页或首页', '手动', 'P0'),

    # ========== DEALS / HOME 模块 ==========
    # 首页
    ('Deals', '首页', '首页正常加载', '已打开App',
     '查看首页',
     '显示城市选择、分类图标、Featured Deals、Deal 列表', '手动', 'P0'),
    ('Deals', '首页', '城市选择（三级菜单）', '在首页',
     '点击城市图标，选择州 → Metro → 城市',
     '城市切换后 Deal 列表刷新为所选城市的数据', '手动', 'P0'),
    ('Deals', '首页', 'Near Me 模式切换', '在首页',
     '点击 "Near Me" 切换开关',
     '切换到 GPS 搜索模式，显示距离信息', '手动', 'P0'),
    ('Deals', '首页', 'GPS 权限请求', '首次使用 Near Me',
     '点击 Near Me，系统弹出定位权限请求',
     '允许后显示附近 deals + 距离；拒绝后显示权限提示条', '手动', 'P0'),
    ('Deals', '首页', '分类筛选', '在首页',
     '点击 "BBQ" 分类图标',
     'Deal 列表只显示 BBQ 分类，图标高亮；再次点击取消筛选', '手动', 'P0'),
    ('Deals', '首页', 'Featured Deals 水平滚动', '首页有展示券',
     '左右滑动 Featured Deals 区域',
     '展示券可水平滚动浏览', '手动', 'P1'),
    ('Deals', '首页', '下拉刷新', '在首页',
     '下拉刷新首页',
     'Featured Deals 和列表数据重新加载', '手动', 'P1'),
    ('Deals', '首页', '点击 Deal 卡片进入详情', '首页有 Deal 列表',
     '点击任意 Deal 卡片',
     '跳转到 DealDetailScreen', '手动', 'P0'),

    # 搜索
    ('Deals', '搜索', '搜索页初始状态', '在首页',
     '点击搜索栏进入搜索页',
     '显示热门标签（BBQ/Sushi/Hot Pot 等）+ 搜索历史', '手动', 'P0'),
    ('Deals', '搜索', '关键词搜索', '在搜索页',
     '输入 "sushi" 并提交',
     '显示 sushi 相关的 deal 结果列表', '手动', 'P0'),
    ('Deals', '搜索', '实时搜索建议', '在搜索页',
     '输入 2 个以上字符（如 "co"）',
     '300ms 后显示建议列表（最多 8 条）', '手动', 'P1'),
    ('Deals', '搜索', '点击热门标签搜索', '在搜索页 Idle 状态',
     '点击 "Coffee" 标签',
     '直接提交搜索，显示 Coffee 结果', '手动', 'P1'),
    ('Deals', '搜索', '搜索历史管理', '有搜索历史',
     '1. 搜索 "pizza"\n2. 返回搜索页查看历史\n3. 点击 × 删除某条\n4. 点击 "Clear All" 清空',
     '历史记录正确显示、删除、清空', '手动', 'P1'),
    ('Deals', '搜索', '过滤条件筛选', '搜索结果页',
     '点击 Filter → 选择分类 "BBQ" + 价格 $10-$50 + 评分 4+',
     '结果按条件过滤', '手动', 'P1'),
    ('Deals', '搜索', '排序功能', '搜索结果页',
     '点击 Sort → 选择 "Price: Low to High"',
     '结果按价格升序排列', '手动', 'P1'),
    ('Deals', '搜索', '无搜索结果', '在搜索页',
     '搜索 "xyznotexist123"',
     '显示空状态图标 + "No results found" 提示', '手动', 'P1'),

    # Deal 详情
    ('Deals', 'Deal 详情', '详情页完整展示', '有可用 Deal',
     '从首页点击 Deal 进入详情',
     '展示图片画廊、价格、标题、销量、商家信息、评价区域', '手动', 'P0'),
    ('Deals', 'Deal 详情', '图片画廊左右滑动', '在详情页',
     '左右滑动图片区域',
     '图片可水平翻页，指示器同步更新', '手动', 'P0'),
    ('Deals', 'Deal 详情', '收藏 Deal', '已登录',
     '点击心形收藏按钮',
     '按钮变为实心红色，deal 加入收藏列表', '手动', 'P0'),
    ('Deals', 'Deal 详情', '取消收藏', '已收藏的 Deal',
     '再次点击心形按钮',
     '按钮恢复空心，deal 从收藏列表移除', '手动', 'P1'),
    ('Deals', 'Deal 详情', '未登录收藏', '未登录状态',
     '点击心形收藏按钮',
     '跳转登录页，登录后可继续操作', '手动', 'P1'),
    ('Deals', 'Deal 详情', '分享 Deal', '在详情页',
     '点击分享按钮',
     '弹出系统分享面板', '手动', 'P1'),
    ('Deals', 'Deal 详情', '选项组选择（几选几）', '有选项组的 Deal',
     '在选项组区域选择/取消选项',
     '选项状态正确更新，满足最少/最多选择数量要求', '手动', 'P0'),
    ('Deals', 'Deal 详情', '查看商家信息', '在详情页',
     '向下滚动到商家信息区域',
     '显示营业时间、电话、地址', '手动', 'P1'),
    ('Deals', 'Deal 详情', '查看评价列表', '在详情页',
     '滚动到评价区域',
     '显示评分分布 + 评价列表（含 5 维度评分）', '手动', 'P1'),
    ('Deals', 'Deal 详情', '点击 Buy Now', '在详情页',
     '点击底部 "Buy Now" 按钮',
     '跳转到 CheckoutScreen', '手动', 'P0'),
    ('Deals', 'Deal 详情', '浏览历史自动记录', '在详情页',
     '查看一个 Deal 后去"最近浏览"',
     '该 Deal 出现在浏览历史列表最前面', '手动', 'P1'),
    ('Deals', 'Deal 详情', '多店通用 Deal 展示适用门店', '品牌 Deal',
     '查看 applicable stores 区域',
     '列出所有适用门店的名称和地址', '手动', 'P1'),

    # 收藏
    ('Deals', '收藏', '我的收藏页 - Deals Tab', '有收藏的 Deals',
     '进入 My Collection → Deals Tab',
     '显示收藏的 Deal 列表', '手动', 'P0'),
    ('Deals', '收藏', '我的收藏页 - Stores Tab', '有收藏的商家',
     '进入 My Collection → Stores Tab',
     '显示收藏的商家列表', '手动', 'P1'),
    ('Deals', '收藏', '收藏为空状态', '无收藏',
     '进入 My Collection',
     '显示空状态提示 + "Explore" 按钮', '手动', 'P1'),

    # 浏览历史
    ('Deals', '浏览历史', '查看浏览历史', '浏览过若干 Deals',
     '进入 History 页',
     '按时间倒序显示浏览过的 Deals', '手动', 'P1'),
    ('Deals', '浏览历史', '清空浏览历史', '有浏览历史',
     '点击 "Clear" 按钮',
     '所有历史记录被清除', '手动', 'P2'),

    # ========== MERCHANT 模块 ==========
    ('Merchant', '商家详情', '商家详情页完整展示', '有可用商家',
     '从首页/搜索进入商家详情',
     '展示头图、Logo、名称、评分、距离、4 Tab（Deals/About/Menu/Reviews）', '手动', 'P0'),
    ('Merchant', '商家详情', 'Deals Tab 展示', '商家有活跃 Deals',
     '在商家详情页查看 Deals Tab',
     '列出该商家所有 Deals，支持分类筛选', '手动', 'P0'),
    ('Merchant', '商家详情', 'About Tab 展示', '在商家详情页',
     '切换到 About Tab',
     '显示地址、电话、营业时间、停车、WiFi 等信息', '手动', 'P1'),
    ('Merchant', '商家详情', 'Menu Tab 展示', '商家有菜单',
     '切换到 Menu Tab',
     '按分类展示菜单项（图片+名称+价格）', '手动', 'P1'),
    ('Merchant', '商家详情', 'Reviews Tab 展示', '商家有评价',
     '切换到 Reviews Tab',
     '显示评分统计图表 + 评价列表', '手动', 'P1'),
    ('Merchant', '商家详情', '收藏商家', '已登录',
     '点击收藏按钮',
     '商家加入收藏列表', '手动', 'P1'),
    ('Merchant', '商家详情', '查看商家相册', '商家有照片',
     '点击查看更多照片',
     '进入 PhotoGalleryScreen，双列网格展示', '手动', 'P2'),
    ('Merchant', '品牌聚合页', '品牌页面展示', '有品牌门店',
     '进入品牌详情页',
     '展示品牌信息 + 旗下所有门店列表', '手动', 'P1'),

    # ========== CART 模块 ==========
    ('Cart', '购物车', '添加 Deal 到购物车', '已登录，在 Deal 详情页',
     '点击 "Add to Cart"',
     '购物车 badge 数字+1，进入购物车可见该项', '手动', 'P0'),
    ('Cart', '购物车', '购物车列表展示', '购物车有商品',
     '点击底部 Cart Tab',
     '按 Deal 分组显示，每组显示数量、单价、商家', '手动', 'P0'),
    ('Cart', '购物车', '增加数量', '购物车有商品',
     '点击 "+" 按钮',
     '数量+1，总价同步更新', '手动', 'P0'),
    ('Cart', '购物车', '减少数量', '购物车某 Deal 数量 > 1',
     '点击 "-" 按钮',
     '数量-1，总价同步更新', '手动', 'P0'),
    ('Cart', '购物车', '限购检查', '购物车某 Deal 有限购',
     '增加数量超过 max_per_account（含已购买的）',
     '提示限购，无法继续增加', '手动', 'P0'),
    ('Cart', '购物车', '滑动删除', '购物车有商品',
     '向左滑动某 Deal 组',
     '该 Deal 组被删除', '手动', 'P1'),
    ('Cart', '购物车', '全选/取消全选', '购物车有多个 Deal',
     '点击全选按钮',
     '所有 Deal 被选中/取消选中，价格同步更新', '手动', 'P1'),
    ('Cart', '购物车', '单选 Deal', '购物车有多个 Deal',
     '勾选/取消单个 Deal',
     '只影响对应 Deal，总价同步更新', '手动', 'P1'),
    ('Cart', '购物车', 'Service Fee 计算', '购物车有商品',
     '查看底部价格明细',
     'Service Fee = $0.99 × 选中券总数', '手动', 'P0'),
    ('Cart', '购物车', '空购物车状态', '购物车为空',
     '点击 Cart Tab',
     '显示空状态 UI + "Go Shopping" 入口', '手动', 'P1'),
    ('Cart', '购物车', '点击 Checkout', '购物车有选中商品',
     '点击 Checkout 按钮',
     '跳转到 CheckoutScreen（购物车模式），携带选中的 items', '手动', 'P0'),

    # ========== CHECKOUT 模块 ==========
    ('Checkout', '支付', '单 Deal 快速购买', '在 Deal 详情页',
     '点击 Buy Now → 选择支付方式 → 完成支付',
     '支付成功，跳转到 OrderSuccessScreen', '手动', 'P0'),
    ('Checkout', '支付', '购物车多 Deal 结账', '购物车有选中商品',
     '从购物车 Checkout → 选择支付方式 → 完成支付',
     '支付成功，所有 Deal 创建订单', '手动', 'P0'),
    ('Checkout', '支付', '信用卡支付', '在结账页',
     '选择 Card → 输入卡号/有效期/CVV → 填写账单地址 → 支付',
     'Stripe 处理成功，订单创建', '手动', 'P0'),
    ('Checkout', '支付', 'Apple Pay 支付（iOS）', 'iOS 设备',
     '选择 Apple Pay → 验证 Face ID/Touch ID',
     '支付成功', '手动', 'P0'),
    ('Checkout', '支付', 'Google Pay 支付（Android）', 'Android 设备',
     '选择 Google Pay → 确认支付',
     '支付成功', '手动', 'P0'),
    ('Checkout', '支付', '使用已保存卡片', '有已保存的卡片',
     '选择已保存的卡 → 确认支付',
     '使用已保存卡片直接支付', '手动', 'P0'),
    ('Checkout', '支付', 'Store Credit 全额抵扣', '余额足够',
     '勾选 Use Store Credit → 余额≥总价',
     '跳过 Stripe 弹窗，直接创建订单', '手动', 'P0'),
    ('Checkout', '支付', 'Store Credit 部分抵扣', '余额不足全额覆盖',
     '勾选 Use Store Credit → 余额 < 总价',
     '抵扣后剩余金额通过 Stripe 支付', '手动', 'P1'),
    ('Checkout', '支付', '优惠码使用', '有有效优惠码',
     '输入优惠码 → 点击验证 → 完成支付',
     '折扣正确应用，总价扣减', '手动', 'P1'),
    ('Checkout', '支付', '无效优惠码', '在结账页',
     '输入无效优惠码 → 点击验证',
     '显示优惠码无效的错误提示', '手动', 'P1'),
    ('Checkout', '支付', '数量调整（单 Deal 模式）', '在结账页',
     '调整购买数量',
     '价格和 Service Fee 实时更新', '手动', 'P1'),
    ('Checkout', '支付', '账单地址管理', '在结账页',
     '1. 选择已保存地址\n2. 或输入新地址\n3. 可选 "Save as default"',
     '地址正确保存和加载', '手动', 'P1'),
    ('Checkout', '支付', '支付成功页展示', '支付完成',
     '查看 OrderSuccessScreen',
     '显示绿色勾号、订单号、Deal 信息、券数量、支付金额、两个导航按钮', '手动', 'P0'),

    # ========== ORDERS 模块 ==========
    ('Orders', '订单列表', '订单列表展示', '有历史订单',
     '进入 Orders 页面',
     '按时间倒序展示订单，每单显示商家、Deal、数量、状态、金额', '手动', 'P0'),
    ('Orders', '订单列表', '点击订单进入详情', '在订单列表',
     '点击某个订单',
     '跳转到 OrderDetailScreen', '手动', 'P0'),
    ('Orders', '订单详情', '订单详情页展示', '有订单',
     '进入订单详情',
     '显示状态横幅、Deal 摘要、券状态、订单信息、操作按钮', '手动', 'P0'),
    ('Orders', '订单详情', 'QR 码弹窗（多张券）', '订单含多张未使用券',
     '点击 "Show QR Code"',
     'QR 码弹窗支持左右滑动，显示分页指示器', '手动', 'P0'),
    ('Orders', '订单详情', '复制券码', '在 QR 码弹窗',
     '点击复制按钮',
     '券码复制到剪贴板', '手动', 'P1'),

    # ========== COUPONS 模块 ==========
    ('Coupons', '我的券', 'Unused Tab 展示', '有未使用的券',
     '进入 My Coupons → Unused Tab',
     '显示未使用券，即将过期的（7天内）独立展示，其他按商家分组', '手动', 'P0'),
    ('Coupons', '我的券', 'Used Tab 展示', '有已使用的券',
     '切换到 Used Tab',
     '显示已核销的券列表', '手动', 'P0'),
    ('Coupons', '我的券', 'To Review Tab', '有已使用未评价的券',
     '切换到 To Review Tab',
     '显示待评价列表（同 Deal 只显示一次）', '手动', 'P1'),
    ('Coupons', '我的券', 'Expired Tab（自动退款）', '有过期券',
     '切换到 Expired Tab',
     '显示已过期且自动退款的券', '手动', 'P1'),
    ('Coupons', '我的券', 'Refunded Tab', '有主动退款的券',
     '切换到 Refunded Tab',
     '显示主动退款的券', '手动', 'P1'),
    ('Coupons', '我的券', 'Gifted Tab', '有赠送的券',
     '切换到 Gifted Tab',
     '显示已赠送出去的券', '手动', 'P1'),
    ('Coupons', '券详情', '未使用券详情展示', '有未使用的券',
     '点击 Unused 券进入详情',
     '显示 READY TO USE 状态、QR 码、Deal 信息、操作按钮', '手动', 'P0'),
    ('Coupons', '券详情', '已使用券详情', '有已使用的券',
     '点击 Used 券进入详情',
     '显示 USED 状态标识、使用时间、评价入口', '手动', 'P1'),

    # ========== REFUND 模块 ==========
    ('Refund', '退款', '申请退款', '有未使用的券',
     '1. 进入券详情\n2. 点击 "Refund"\n3. 查看退款金额和退回方式\n4. 选择退款原因（可选）\n5. 点击 "Confirm Refund"',
     '退款申请提交成功，状态变为 Processing', '手动', 'P0'),
    ('Refund', '退款', '退款金额展示', '在退款页',
     '查看退款金额',
     '显示退款金额 = 单价 + 税费（不含 service fee）', '手动', 'P0'),
    ('Refund', '退款', '退款方式', '在退款页',
     '查看退回方式',
     '显示原支付方式（Apple Pay 1-3天 / 信用卡 3-5天）', '手动', 'P1'),
    ('Refund', '退款', '已退款券无法再退', '已退款的券',
     '查看已退款券详情',
     '"Refund" 按钮不显示或禁用', '手动', 'P1'),
    ('Refund', '退款', '已使用券无法退款', '已使用的券',
     '查看已使用券详情',
     '退款按钮不显示（可走售后流程）', '手动', 'P1'),
    ('Refund', '自动退款', '过期券自动退款', '券已过期',
     '等待自动退款定时任务执行',
     '券状态变为 Expired Refund，退款到原支付方式', '后端验证', 'P0'),

    # ========== GIFT 模块 ==========
    ('Gift', '赠送', '发送礼品券', '有未使用的券',
     '1. 点击 "Gift" 按钮\n2. 输入收件人邮箱\n3. 输入留言（可选）\n4. 点击 Send Gift',
     '赠送成功，券状态变为 Gifted', '手动', 'P0'),
    ('Gift', '赠送', '通过手机号赠送', '有未使用的券',
     '输入收件人手机号 → Send Gift',
     '赠送成功', '手动', 'P1'),
    ('Gift', '赠送', '撤回赠送', '已赠送未领取',
     '点击 "Recall Gift"',
     '赠送撤回，券恢复为 Unused 状态', '手动', 'P0'),
    ('Gift', '赠送', '领取礼品券', '收到赠送通知',
     '通过 deep link 打开 GiftClaimScreen → 点击领取',
     '领取成功，券出现在自己的 My Coupons 中', '手动', 'P0'),
    ('Gift', '赠送', '已撤回的礼品券领取', '赠送已被撤回',
     '通过 deep link 尝试领取',
     '显示 "已被撤回" 提示', '手动', 'P1'),
    ('Gift', '赠送', '已过期的礼品链接', '链接已过期',
     '通过过期 deep link 尝试领取',
     '显示 "链接已过期" 提示', '手动', 'P1'),

    # ========== REVIEWS 模块 ==========
    ('Reviews', '评价', '写评价', '有已使用未评价的券',
     '1. To Review Tab → 点击 "Review"\n2. 选择总体评分（1-5星）\n3. 可选填子维度评分\n4. 选择 Hashtags\n5. 填写评论文字\n6. 点击提交',
     '评价成功，刷新 Deal 评分，返回上一页', '手动', 'P0'),
    ('Reviews', '评价', '编辑已有评价', '已写过评价',
     '从评价入口进入编辑模式',
     '预填所有已有字段，可修改后重新提交', '手动', 'P1'),
    ('Reviews', '评价', '未选总体评分提交', '在评价页',
     '不选总体评分直接提交',
     '提示总体评分为必填', '手动', 'P1'),
    ('Reviews', '评价', 'Hashtag 选择', '在评价页',
     '选择/取消 Hashtag 标签',
     '标签正确切换选中状态', '手动', 'P2'),

    # ========== PROFILE 模块 ==========
    ('Profile', '个人中心', '个人中心页展示', '已登录',
     '点击 Profile Tab',
     '显示头像、用户名、收藏/历史/券/余额快捷入口、订单状态导航、设置菜单', '手动', 'P0'),
    ('Profile', '个人中心', '游客模式展示', '未登录',
     '点击 Profile Tab',
     '显示 Guest 登录入口', '手动', 'P1'),
    ('Profile', '编辑资料', '修改头像', '已登录',
     '1. 点击编辑资料\n2. 点击头像\n3. 选择相机/相册\n4. 裁剪并保存',
     '头像更新成功，返回后显示新头像', '手动', 'P1'),
    ('Profile', '编辑资料', '修改姓名和简介', '在编辑资料页',
     '修改 Full Name 和 Bio → 保存',
     '信息更新成功', '手动', 'P1'),
    ('Profile', '编辑资料', '简介字数限制', '在编辑资料页',
     '输入超过 150 字的 Bio',
     '显示字数统计，限制不超过 150 字', '手动', 'P2'),
    ('Profile', '修改邮箱', '两步验证修改邮箱', '已登录',
     '1. 进入 Change Email\n2. 输入新邮箱\n3. 发送验证码\n4. 输入收到的验证码\n5. 确认',
     '邮箱修改成功，同步更新 Auth 和 users 表', '手动', 'P0'),
    ('Profile', '修改密码', '修改密码', '已登录',
     '输入当前密码 + 新密码 → 保存',
     '密码修改成功', '手动', 'P1'),
    ('Profile', '支付方式', '查看已保存卡片', '有保存的卡',
     '进入 Payment Methods',
     '列表显示卡片品牌、末四位、过期日期', '手动', 'P1'),
    ('Profile', '支付方式', '添加新卡', '在支付方式页',
     '点击 "Add New Card" → 完成 Stripe 表单',
     '卡片保存成功，出现在列表中', '手动', 'P1'),
    ('Profile', '支付方式', '删除卡片', '有保存的卡',
     '左滑卡片 → 确认删除',
     '卡片从列表移除', '手动', 'P1'),
    ('Profile', '支付方式', '设置默认卡', '有多张卡',
     '点击非默认卡',
     '该卡标记为默认', '手动', 'P2'),
    ('Profile', 'Store Credit', '余额展示', '有 Store Credit',
     '进入 Store Credit 页',
     '显示可用余额 + 交易记录列表', '手动', 'P1'),
    ('Profile', '登出', '登出功能', '已登录',
     '点击 "Sign Out" 按钮',
     '清除登录态，返回首页/欢迎页', '手动', 'P0'),

    # ========== CHAT 模块 ==========
    ('Chat', '会话列表', '会话列表展示', '已登录且有聊天记录',
     '点击 Chat Tab',
     '显示会话列表，Support 固定顶部，其他按时间排序', '手动', 'P0'),
    ('Chat', '会话列表', '未读消息 Badge', '有未读消息',
     '查看会话列表',
     '对应会话显示未读数 Badge', '手动', 'P1'),
    ('Chat', '会话列表', '置顶/取消置顶会话', '在会话列表',
     '左滑会话 → 点击 Pin/Unpin',
     '会话置顶到顶部或取消置顶', '手动', 'P2'),
    ('Chat', '会话列表', '删除会话', '在会话列表',
     '左滑会话 → 点击 Delete → 确认',
     '会话从列表移除', '手动', 'P2'),
    ('Chat', '聊天详情', '发送文字消息', '在聊天详情页',
     '输入文字 → 点击发送',
     '消息出现在聊天列表底部', '手动', 'P0'),
    ('Chat', '聊天详情', '发送图片', '在聊天详情页',
     '点击图片按钮 → 选择图片 → 发送',
     '图片消息出现在聊天列表', '手动', 'P1'),
    ('Chat', '聊天详情', '发送优惠券', '在聊天详情页',
     '点击优惠券按钮 → 选择券 → 发送',
     '优惠券消息卡片出现在聊天列表', '手动', 'P1'),
    ('Chat', '聊天详情', 'Realtime 实时消息接收', '在聊天详情页',
     '对方发送消息',
     '消息实时出现在聊天列表，无需刷新', '手动', 'P0'),
    ('Chat', '聊天详情', '历史消息分页加载', '聊天记录超过 30 条',
     '向上滚动到顶部',
     '自动加载更多历史消息', '手动', 'P1'),
    ('Chat', '聊天详情', '消息已读标记', '进入聊天详情',
     '进入有未读消息的会话',
     '未读消息自动标记为已读', '手动', 'P1'),
    ('Chat', '客服聊天', 'AI 客服对话', '发起客服聊天',
     '进入 Support 会话 → 发送问题',
     'AI 自动回复相关答案', '手动', 'P0'),
    ('Chat', '客服聊天', 'AI 转人工', '客服无法解决问题',
     '连续发送 AI 无法解决的问题',
     'AI 提示转接人工客服，会话状态变更', '手动', 'P1'),
    ('Chat', '好友管理', '搜索用户添加好友', '已登录',
     '1. 进入好友列表\n2. 搜索用户\n3. 发送好友申请',
     '好友申请发送成功', '手动', 'P1'),
    ('Chat', '好友管理', '接受/拒绝好友申请', '有待处理申请',
     '在好友申请页点击接受/拒绝',
     '好友关系建立或申请被拒', '手动', 'P1'),
    ('Chat', '搜索', '搜索会话', '在聊天页',
     '点击搜索 → 输入关键词',
     '搜索匹配的会话和用户', '手动', 'P2'),

    # ========== AFTER-SALES 模块 ==========
    ('After-Sales', '售后', '提交售后申请', '有已使用的券（7天内）',
     '1. 进入券详情\n2. 点击 "After-Sales"\n3. 选择原因\n4. 描述问题\n5. 提交',
     '售后申请提交成功，状态为 Pending', '手动', 'P0'),
    ('After-Sales', '售后', '查看售后进度', '有售后申请',
     '进入售后时间线页面',
     '显示售后处理进度（时间线形式）', '手动', 'P1'),
    ('After-Sales', '售后', '超过 7 天无法申请', '券使用超过 7 天',
     '尝试提交售后',
     '售后入口不显示或提示已超过申请期限', '手动', 'P1'),
    ('After-Sales', '售后', '升级到平台处理', '商家拒绝了售后',
     '点击 "Escalate to Platform"',
     '售后升级为平台处理', '手动', 'P1'),

    # ========== NAVIGATION / GENERAL 模块 ==========
    ('Navigation', '底部导航', '4 Tab 切换', '在任意页面',
     '依次点击 Deals / Chat / Cart / Profile Tab',
     '4 个 Tab 页面正确切换', '手动', 'P0'),
    ('Navigation', '底部导航', 'Cart Tab Badge', '购物车有商品',
     '查看底部导航',
     'Cart Tab 显示商品数量 Badge', '手动', 'P1'),
    ('Navigation', '路由守卫', '未登录访问受保护页面', '未登录',
     '尝试直接访问 /coupons、/orders、/checkout 等',
     '重定向到登录页，登录后返回原页面', '手动', 'P0'),
    ('Navigation', '路由守卫', '未登录可访问公开页面', '未登录',
     '直接访问 /home、/deals/:id、/search',
     '可正常浏览，无需登录', '手动', 'P0'),
    ('Navigation', '邮箱验证横幅', '未验证邮箱横幅', '已登录但邮箱未验证',
     '查看底部导航栏上方',
     '显示橙色横幅提示验证邮箱 + Resend 按钮', '手动', 'P1'),
    ('Navigation', 'Realtime', '订单/券状态实时更新', '已登录',
     '商家扫码核销券',
     'App 自动更新券状态为 Used，无需手动刷新', '手动+后端', 'P0'),
    ('Navigation', 'Deep Link', '礼品领取 Deep Link', '收到礼品链接',
     '点击 crunchyplum://gift?token=xxx',
     'App 打开并跳转到 GiftClaimScreen', '手动', 'P1'),

    # ========== EDGE FUNCTION / BACKEND 模块 ==========
    ('Backend', '支付流程', 'create-payment-intent 价格防篡改', '无',
     '前端传入价格与 DB 不一致',
     '后端检测价格不一致，拒绝创建支付意图', '后端验证', 'P0'),
    ('Backend', '支付流程', '限购校验', '无',
     '购买数量超过 max_per_account',
     '后端拒绝，返回限购错误', '后端验证', 'P0'),
    ('Backend', '支付流程', '幂等性 — 同一 PI 不重复创建', '无',
     '相同 payment_intent_id 重复调用 create-order-v3',
     '后端返回已存在的订单，不重复创建', '后端验证', 'P0'),
    ('Backend', 'Webhook', 'Stripe Webhook 正常处理', '支付完成',
     'Stripe 发送 payment_intent.succeeded 事件',
     '后端正确更新订单状态、发送确认邮件', '后端验证', 'P0'),
    ('Backend', 'Webhook', 'charge.refunded 事件处理', '退款完成',
     'Stripe 发送 charge.refunded 事件',
     '后端更新券状态为 refunded，发送退款通知邮件', '后端验证', 'P0'),
    ('Backend', '自动退款', 'auto-refund-expired 定时任务', '有过期未退款的券',
     '定时任务触发执行',
     '过期券自动退款（退 unit_price + tax，不退 service_fee）', '后端验证', 'P0'),
    ('Backend', '自动扣款', 'auto-capture-preauth 定时任务', '有 6 天前的预授权',
     '定时任务触发执行',
     '自动 capture 预授权金额', '后端验证', 'P1'),
    ('Backend', '商家核销', 'merchant-scan 扫码核销', '商家扫码',
     '扫描有效券的 QR 码',
     '券状态变 used，记录核销门店和时间', '后端验证', 'P0'),
    ('Backend', '客服 AI', 'support-chat AI 回复', '发送客服消息',
     '调用 support-chat Edge Function',
     'Claude 生成回复，检测转人工关键词', '后端验证', 'P1'),
    ('Backend', '推荐系统', 'get-recommendations 推荐', '有用户标签',
     '调用推荐接口',
     '返回个性化推荐列表', '后端验证', 'P2'),
    ('Backend', '邮件通知', '订单确认邮件', '支付成功',
     '创建订单后检查邮箱',
     '收到订单确认邮件（C2）', '后端验证', 'P1'),
    ('Backend', '邮件通知', '自动退款通知邮件', '自动退款触发',
     '自动退款后检查邮箱',
     '收到自动退款通知邮件（C5）', '后端验证', 'P1'),
]


def create_sheet(ws, title, cases):
    """创建一个测试用例 Sheet"""
    ws.title = title

    # 写入表头
    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 写入数据
    row = 2
    case_num = 1
    current_module = ''

    for case in cases:
        module, sub_module, scenario, precondition, steps, expected, method, priority = case

        ws.cell(row=row, column=1, value=f'TC-{case_num:03d}')
        ws.cell(row=row, column=2, value=module)
        ws.cell(row=row, column=3, value=sub_module)
        ws.cell(row=row, column=4, value=scenario)
        ws.cell(row=row, column=5, value=precondition)
        ws.cell(row=row, column=6, value=steps)
        ws.cell(row=row, column=7, value=expected)
        ws.cell(row=row, column=8, value=method)
        ws.cell(row=row, column=9, value=priority)
        ws.cell(row=row, column=10, value='')  # 测试结果（待填）
        ws.cell(row=row, column=11, value='')  # 备注（待填）

        # 样式
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.font = normal_font
            cell.alignment = wrap_alignment
            cell.border = thin_border

        # 模块行高亮
        if module != current_module:
            current_module = module
            ws.cell(row=row, column=2).fill = module_fill
            ws.cell(row=row, column=2).font = module_font

        # 优先级颜色
        priority_cell = ws.cell(row=row, column=9)
        priority_cell.alignment = Alignment(horizontal='center', vertical='top')
        if priority == 'P0':
            priority_cell.font = Font(name='Arial', size=10, bold=True, color='C00000')
        elif priority == 'P1':
            priority_cell.font = Font(name='Arial', size=10, color='E07000')

        row += 1
        case_num += 1

    # 设置行高
    for r in range(2, row):
        ws.row_dimensions[r].height = 45

    ws.row_dimensions[1].height = 25


# 创建主Sheet
ws = wb.active
create_sheet(ws, '用户端测试用例', test_cases)

# ===== 创建统计 Sheet =====
ws_stats = wb.create_sheet('测试统计')

# 统计数据
modules = {}
priorities = {'P0': 0, 'P1': 0, 'P2': 0}
methods = {}

for case in test_cases:
    module = case[0]
    priority = case[7]
    method = case[6]

    modules[module] = modules.get(module, 0) + 1
    priorities[priority] = priorities.get(priority, 0) + 1
    methods[method] = methods.get(method, 0) + 1

# 模块统计表
ws_stats.cell(row=1, column=1, value='模块统计').font = Font(bold=True, size=12)
ws_stats.cell(row=2, column=1, value='模块').font = Font(bold=True)
ws_stats.cell(row=2, column=2, value='用例数').font = Font(bold=True)
ws_stats.cell(row=2, column=3, value='占比').font = Font(bold=True)

total = len(test_cases)
for i, (module, count) in enumerate(sorted(modules.items(), key=lambda x: -x[1]), 3):
    ws_stats.cell(row=i, column=1, value=module)
    ws_stats.cell(row=i, column=2, value=count)
    ws_stats.cell(row=i, column=3, value=f'{count/total*100:.1f}%')

# 优先级统计
stats_row = len(modules) + 5
ws_stats.cell(row=stats_row, column=1, value='优先级统计').font = Font(bold=True, size=12)
ws_stats.cell(row=stats_row+1, column=1, value='优先级').font = Font(bold=True)
ws_stats.cell(row=stats_row+1, column=2, value='用例数').font = Font(bold=True)
ws_stats.cell(row=stats_row+1, column=3, value='说明').font = Font(bold=True)

desc = {'P0': '核心功能，必须通过', 'P1': '重要功能，优先修复', 'P2': '次要功能，可延后'}
for i, (p, count) in enumerate(sorted(priorities.items()), stats_row+2):
    ws_stats.cell(row=i, column=1, value=p)
    ws_stats.cell(row=i, column=2, value=count)
    ws_stats.cell(row=i, column=3, value=desc.get(p, ''))

# 测试方式统计
method_row = stats_row + 7
ws_stats.cell(row=method_row, column=1, value='测试方式统计').font = Font(bold=True, size=12)
ws_stats.cell(row=method_row+1, column=1, value='方式').font = Font(bold=True)
ws_stats.cell(row=method_row+1, column=2, value='用例数').font = Font(bold=True)
for i, (m, count) in enumerate(sorted(methods.items()), method_row+2):
    ws_stats.cell(row=i, column=1, value=m)
    ws_stats.cell(row=i, column=2, value=count)

# 列宽
ws_stats.column_dimensions['A'].width = 18
ws_stats.column_dimensions['B'].width = 10
ws_stats.column_dimensions['C'].width = 25

# 保存
output_path = '/Users/howardshansmac/github/coupon app/coupon-app/DealJoy_测试用例.xlsx'
wb.save(output_path)
print(f'✅ 测试用例已生成：{output_path}')
print(f'📊 总计 {len(test_cases)} 个测试用例')
print(f'   P0: {priorities["P0"]} | P1: {priorities["P1"]} | P2: {priorities["P2"]}')
print(f'   模块分布: {dict(sorted(modules.items(), key=lambda x: -x[1]))}')
