#!/usr/bin/env python3
"""
从 DealJoy 需求清单 Excel 中提取指定模块的数据。
用法: python3 scripts/read_excel.py "1.用户认证系统"
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    os.system("pip install openpyxl --break-system-packages -q")
    import openpyxl


def read_module(excel_path: str, module_name: str) -> str:
    """提取指定模块的所有行数据"""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    rows = []
    capturing = False

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        func_sys = row_dict.get("功能系统", "")

        if func_sys and module_name in str(func_sys):
            capturing = True
        if capturing and func_sys and module_name not in str(func_sys):
            if func_sys.strip():
                break
        if capturing:
            rows.append(row_dict)

    wb.close()

    if not rows:
        print(f"❌ 未找到模块: {module_name}", file=sys.stderr)
        sys.exit(1)

    # 输出格式化文本
    output = [f"## 模块: {module_name}", f"共 {len(rows)} 行需求数据\n"]
    for i, r in enumerate(rows):
        non_empty = {k: v for k, v in r.items() if v is not None}
        output.append(f"[行{i+1}] {json.dumps(non_empty, ensure_ascii=False)}")

    return "\n".join(output)


def find_excel():
    """查找需求清单 Excel 文件"""
    candidates = [
        "requirements/DealJoy_V1_详细需求清单_v3.xlsx",
        "DealJoy_V1_详细需求清单_v3.xlsx",
        "requirements/DealJoy_V1_详细需求清单_v2.xlsx",
    ]
    for path in candidates:
        if os.path.exists(path):
            return path

    # 递归搜索
    for root, dirs, files in os.walk("."):
        for f in files:
            if "需求清单" in f and f.endswith(".xlsx"):
                return os.path.join(root, f)

    return None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("用法: python3 scripts/read_excel.py <模块名>")
        print("示例: python3 scripts/read_excel.py '1.用户认证系统'")
        sys.exit(1)

    module_name = sys.argv[1]
    excel_path = sys.argv[2] if len(sys.argv) > 2 else find_excel()

    if not excel_path:
        print("❌ 找不到需求清单 Excel 文件", file=sys.stderr)
        print("请将文件放在 requirements/ 目录下", file=sys.stderr)
        sys.exit(1)

    result = read_module(excel_path, module_name)
    print(result)
