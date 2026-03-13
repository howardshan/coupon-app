#!/usr/bin/env python3
"""
DealJoy 三端测试计划生成器 v2
输出: test_plan.xlsx（5 个 Sheet）
修复：三端联动列错位、补充 Key 字段、增加边界/错误处理测试项
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── 颜色常量 ─────────────────────────────────────────────
HEADER_BG   = "1F3864"   # 深蓝
HEADER_FG   = "FFFFFF"   # 白字
HIGH_BG     = "FFE0E0"   # 高优先级行 浅红
CLIENT_BG   = "E3F2FD"   # 客户端列 浅蓝
MERCHANT_BG = "E8F5E9"   # 商家端列 浅绿
ADMIN_BG    = "FFF3E0"   # 后台列 浅橙
ALT_ROW_BG  = "F5F7FA"   # 交替行

def make_header_fill():
    return PatternFill("solid", fgColor=HEADER_BG)

def make_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def header_font():
    return Font(bold=True, color=HEADER_FG, size=11)

def normal_font(bold=False):
    return Font(bold=bold, size=10)

def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def write_header(ws, headers, col_widths=None):
    """写表头行（第1行），设置样式"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = make_header_fill()
        cell.font = header_font()
        cell.alignment = center()
        cell.border = thin_border()
    if col_widths:
        for col_idx, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = w

def write_row(ws, row_idx, values, fill=None, bold=False):
    for col_idx, v in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=v)
        if fill:
            cell.fill = fill
        cell.font = normal_font(bold=bold)
        cell.alignment = left()
        cell.border = thin_border()

def freeze_and_filter(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# ─────────────────────────────────────────────────────────
# Sheet 1：总览
# ─────────────────────────────────────────────────────────
def build_overview(wb):
    ws = wb.create_sheet("📊 总览")
    ws.row_dimensions[1].height = 28

    headers = ["测试编号", "测试场景名称", "涉及端", "测试项总数", "通过数", "失败数", "完成率"]
    widths  = [12, 42, 20, 12, 10, 10, 12]
    write_header(ws, headers, widths)

    # 汇总行数据（与各 Sheet 行数对应）
    data = [
        ("S001", "客户端功能测试汇总",   "客户端",   82, "", "", ""),
        ("S002", "商家端功能测试汇总",   "商家端",  109, "", "", ""),
        ("S003", "后台管理功能测试汇总", "后台",     60, "", "", ""),
        ("S004", "三端联动测试汇总",    "多端联动",  20, "", "", ""),
    ]

    for r_idx, row in enumerate(data, 2):
        fill = make_fill(ALT_ROW_BG) if r_idx % 2 == 0 else None
        write_row(ws, r_idx, row, fill=fill)
        col_d = get_column_letter(4)
        col_e = get_column_letter(5)
        g_cell = ws.cell(row=r_idx, column=7)
        g_cell.value = f"=IF({col_d}{r_idx}=0,\"-\",{col_e}{r_idx}/{col_d}{r_idx})"
        g_cell.number_format = "0%"
        g_cell.alignment = center()
        g_cell.border = thin_border()

    freeze_and_filter(ws)
    return ws

# ─────────────────────────────────────────────────────────
# Sheet 2：客户端测试
# ─────────────────────────────────────────────────────────
def build_client(wb):
    ws = wb.create_sheet("📱 客户端")
    ws.row_dimensions[1].height = 28

    # 列: 编号 | 页面名称 | 测试项描述 | 代码字段名/Key | 测试类型 | 优先级 | 测试状态 | 备注
    headers = ["编号", "页面名称", "测试项描述", "代码字段名/Key", "测试类型", "优先级", "测试状态", "备注"]
    widths  = [8, 20, 48, 30, 16, 8, 10, 24]
    write_header(ws, headers, widths)

    rows = [
        # ── 登录 ──
        ("C001", "登录页", "邮箱输入框可正常输入", "_emailCtrl", "UI显示", "高", "", ""),
        ("C002", "登录页", "密码输入框可正常输入（密码隐藏）", "_passwordCtrl", "UI显示", "高", "", ""),
        ("C003", "登录页", "点击 Sign In — 邮箱格式错误时显示验证提示", "_emailCtrl", "边界输入", "高", "", "RFC-5322 格式"),
        ("C004", "登录页", "点击 Sign In — 密码少于8字符时显示验证提示", "_passwordCtrl", "边界输入", "高", "", ""),
        ("C005", "登录页", "有效凭证登录成功后跳转 /home", "_emailCtrl / _passwordCtrl", "功能操作", "高", "", ""),
        ("C006", "登录页", "点击 Forgot password? 跳转至 /auth/forgot-password", "", "功能操作", "高", "", ""),
        ("C007", "登录页", "点击 Continue with Google 发起 Google 登录", "", "功能操作", "高", "", ""),
        ("C008", "登录页", "点击 Sign Up 跳转至 /auth/register", "", "功能操作", "中", "", ""),
        # ── 注册 ──
        ("C009", "注册页", "Username 字段：少于2字符时显示验证错误", "_usernameCtrl", "边界输入", "高", "", "2-30字符限制"),
        ("C010", "注册页", "Username 字段：超过30字符时显示验证错误", "_usernameCtrl", "边界输入", "中", "", ""),
        ("C011", "注册页", "Full Name 字段：为空时显示必填提示", "_fullNameCtrl", "边界输入", "高", "", ""),
        ("C012", "注册页", "Email 字段：无效格式时显示验证提示", "_emailCtrl", "边界输入", "高", "", ""),
        ("C013", "注册页", "Password 强度不足时指示器实时反馈（4条规则）", "_passwordCtrl", "UI显示", "高", "", "大写/小写/数字/8字符"),
        ("C014", "注册页", "Confirm Password 不匹配时显示错误", "_confirmPasswordCtrl", "边界输入", "高", "", ""),
        ("C015", "注册页", "未勾选 Terms of Service 时点击 Create Account 被阻止", "", "边界输入", "高", "", ""),
        ("C016", "注册页", "填写所有字段并勾选条款后，成功提交注册（显示邮件确认提示）", "_emailCtrl / _passwordCtrl / _usernameCtrl", "功能操作", "高", "", ""),
        # ── 忘记密码 ──
        ("C017", "忘记密码页", "输入有效邮箱后点击 Send Reset Link，显示成功状态", "_emailCtrl", "功能操作", "高", "", ""),
        ("C018", "忘记密码页", "发送后出现60秒倒计时重发按钮（Resend Link）", "", "UI显示", "中", "", ""),
        ("C019", "忘记密码页", "点击 Back to Sign In 跳转至登录页", "", "功能操作", "中", "", ""),
        # ── 重置密码 ──
        ("C020", "重置密码页", "密码强度实时显示，符合4条规则后变绿", "_passwordCtrl", "UI显示", "中", "", ""),
        ("C021", "重置密码页", "Confirm Password 不匹配时显示错误", "_confirmPasswordCtrl", "边界输入", "高", "", ""),
        ("C022", "重置密码页", "密码重置成功后 3s 倒计时跳转 /auth/login", "", "功能操作", "高", "", ""),
        # ── 首页 ──
        ("C023", "首页", "Deal 列表正确加载并显示图片、标题、价格", "", "UI显示", "高", "", ""),
        ("C024", "首页", "搜索框输入关键词，防抖 300ms 后显示搜索建议", "_searchCtrl", "功能操作", "高", "", ""),
        ("C025", "首页", "城市/区域三级选择（州→Metro→城市），切换后列表刷新", "", "功能操作", "高", "", ""),
        ("C026", "首页", "点击 Deal 卡片跳转至 /deals/:id", "", "功能操作", "高", "", ""),
        ("C027", "首页", "收藏按钮点击后 Deal 加入收藏列表", "", "功能操作", "中", "", ""),
        # ── 搜索页 ──
        ("C028", "搜索页", "热门标签点击（BBQ、Sushi等）触发搜索", "", "功能操作", "中", "", ""),
        ("C029", "搜索页", "输入关键词后显示搜索结果列表", "_searchCtrl", "功能操作", "高", "", ""),
        ("C030", "搜索页", "无结果时显示空状态提示", "", "UI显示", "中", "", ""),
        ("C031", "搜索页", "搜索历史记录可正常显示和清除", "", "功能操作", "低", "", ""),
        # ── Deal 详情 ──
        ("C032", "Deal详情页", "页面完整显示：标题、折扣价、原价、描述、使用规则、门店信息", "", "UI显示", "高", "", ""),
        ("C033", "Deal详情页", "点击 Checkout 跳转至 /checkout/:dealId", "", "功能操作", "高", "", ""),
        ("C034", "Deal详情页", "多门店 Deal 显示门店数量和门店列表", "", "UI显示", "中", "", "适用门店 count"),
        ("C035", "Deal详情页", "收藏按钮状态正确（已收藏/未收藏）", "", "UI显示", "中", "", ""),
        # ── 支付页 ──
        ("C036", "支付页", "显示正确的 Deal 名称、单价、数量选择器", "", "UI显示", "高", "", ""),
        ("C037", "支付页", "数量 + / - 按钮可正常加减（最小1，最大库存或限购数）", "", "边界输入", "高", "", ""),
        ("C038", "支付页", "输入有效优惠码后点击 Apply，折扣正确计算", "_couponCtrl", "功能操作", "中", "", ""),
        ("C039", "支付页", "输入无效优惠码显示错误提示", "_couponCtrl", "错误处理", "中", "", ""),
        ("C040", "支付页", "支付方式切换（Apple Pay/Google Pay/信用卡）", "", "UI显示", "高", "", ""),
        ("C041", "支付页", "点击 Confirm Payment 发起支付，成功后跳转 /order-success/:orderId", "", "功能操作", "高", "", ""),
        ("C042", "支付页", "支付失败时显示错误提示，不跳转", "", "错误处理", "高", "", ""),
        # ── 订单 ──
        ("C043", "订单列表页", "正确显示所有订单（订单号、Deal名、金额、状态）", "", "UI显示", "高", "", ""),
        ("C044", "订单列表页", "点击已使用优惠券订单跳转至 /coupon/:couponId", "", "功能操作", "高", "", ""),
        ("C045", "订单列表页", "点击退款入口跳转至 /refund/:orderId", "", "功能操作", "高", "", ""),
        # ── 优惠券 ──
        ("C046", "优惠券页", "4个Tab正确筛选（Unused/Used/Expired/Refunded）", "", "UI显示", "高", "", ""),
        ("C047", "优惠券页", "Unused 优惠券显示有效 QR 码", "", "UI显示", "高", "", ""),
        ("C048", "优惠券页", "空状态时显示 Browse Deals 按钮，点击跳转 /home", "", "UI显示", "中", "", ""),
        # ── 个人中心 ──
        ("C049", "个人中心", "显示用户昵称/邮箱", "", "UI显示", "中", "", ""),
        ("C050", "个人中心", "Collection 入口跳转 /collection", "", "功能操作", "中", "", ""),
        ("C051", "个人中心", "History 入口跳转 /history", "", "功能操作", "中", "", ""),
        ("C052", "个人中心", "Coupons 入口跳转 /coupons", "", "功能操作", "中", "", ""),
        ("C053", "个人中心", "点击 Log Out 退出登录，跳转 /auth/login", "", "功能操作", "高", "", ""),
        # ── 评价 ──
        ("C054", "评价页", "1-5星评分可正常选择，初始4星", "", "UI显示", "中", "", ""),
        ("C055", "评价页", "评论文字框可输入内容", "_commentCtrl", "UI显示", "中", "", ""),
        ("C056", "评价页", "点击 Submit Review 提交成功后跳转回Deal详情", "_commentCtrl", "功能操作", "中", "", ""),
        # ── 退款 ──
        ("C057", "退款申请页", "显示订单信息和退款申请表单", "", "UI显示", "高", "", ""),
        ("C058", "退款申请页", "成功提交退款申请，订单状态变为 refund_requested", "", "功能操作", "高", "", ""),
        # ── 商家详情（用户视角）──
        ("C059", "商家详情页", "显示商家名称、分类、地址、营业时间、评分", "", "UI显示", "中", "", ""),
        ("C060", "商家详情页", "显示该商家所有活跃 Deal", "", "UI显示", "中", "", ""),
        ("C061", "商家详情页", "点击照片查看全屏相册", "", "功能操作", "低", "", ""),
        # ── 收藏 ──
        ("C062", "收藏页", "显示所有已收藏 Deal 列表", "", "UI显示", "中", "", ""),
        ("C063", "收藏页", "取消收藏后从列表移除", "", "功能操作", "中", "", ""),
        # ── 权限/路由 ──
        ("C064", "路由守卫", "未登录用户访问 /home 重定向至 /auth/login", "", "功能操作", "高", "", ""),
        ("C065", "路由守卫", "已登录用户访问 /auth/login 直接跳转 /home", "", "功能操作", "高", "", ""),
        ("C066", "路由守卫", "网络错误时页面显示友好错误提示和重试按钮", "", "错误处理", "中", "", ""),
        # ── 边界输入 & 错误处理（C067-C082）──
        ("C067", "搜索页", "搜索输入超过200字符不崩溃，正常截断或提示", "_searchCtrl", "边界输入", "中", "", "超长字符串"),
        ("C068", "支付页", "数量超过单次限购上限时显示上限提示，无法继续增加", "", "边界输入", "高", "", "max_per_person 字段"),
        ("C069", "支付页", "数量减到0时 Confirm Payment 按钮变灰禁用", "", "边界输入", "高", "", ""),
        ("C070", "支付页", "网络断开后点击 Confirm Payment 显示错误提示而非崩溃", "", "错误处理", "高", "", "离线支付场景"),
        ("C071", "注册页", "已存在邮箱注册时显示 Email already in use 提示", "_emailCtrl", "错误处理", "高", "", "Supabase Auth 错误码"),
        ("C072", "登录页", "错误密码登录时显示 Invalid credentials 提示，不泄露具体信息", "_passwordCtrl", "错误处理", "高", "", "安全性验证"),
        ("C073", "退款申请页", "订单状态为 used 且超过退款窗口期时，退款按钮禁用或不可见", "", "边界输入", "高", "", ""),
        ("C074", "Deal详情页", "Deal 库存为 0 时 Checkout 按钮显示 Sold Out 并禁用", "", "边界输入", "高", "", "stock=0 场景"),
        ("C075", "Deal详情页", "Deal 已过期（validity_end < now）时显示 Expired 标签", "", "边界输入", "高", "", ""),
        ("C076", "支付页", "Stripe Token 创建失败（卡号无效）时显示明确错误文案", "", "错误处理", "高", "", "Stripe 错误码处理"),
        ("C077", "优惠券页", "QR 码图片加载失败时显示占位符，不显示空白", "", "错误处理", "中", "", ""),
        ("C078", "首页", "定位权限被拒绝时降级为全国 Deal 展示，不崩溃", "", "错误处理", "中", "", ""),
        ("C079", "搜索页", "输入 <script>alert(1)</script> 等特殊字符不触发 XSS，显示为普通文本", "_searchCtrl", "错误处理", "高", "", "安全性"),
        ("C080", "个人中心", "网络断开时页面显示上次缓存数据而非白屏", "", "错误处理", "中", "", ""),
        ("C081", "登录页", "邮箱超过 254 字符时显示格式验证错误", "_emailCtrl", "边界输入", "中", "", "RFC 5321 限制"),
        ("C082", "注册页", "密码仅包含空格时视为无效密码，显示验证错误", "_passwordCtrl", "边界输入", "高", "", ""),
    ]

    for r_idx, row in enumerate(rows, 2):
        priority = row[5]
        fill = make_fill(HIGH_BG) if priority == "高" else (make_fill(ALT_ROW_BG) if r_idx % 2 == 0 else None)
        write_row(ws, r_idx, row, fill=fill)

    freeze_and_filter(ws)
    return ws

# ─────────────────────────────────────────────────────────
# Sheet 3：商家端测试
# ─────────────────────────────────────────────────────────
def build_merchant(wb):
    ws = wb.create_sheet("🏪 商家端")
    ws.row_dimensions[1].height = 28

    headers = ["编号", "页面名称", "测试项描述", "代码字段名/Key", "测试类型", "优先级", "测试状态", "备注"]
    widths  = [8, 24, 48, 34, 16, 8, 10, 24]
    write_header(ws, headers, widths)

    rows = [
        # ── 登录 ──
        ("M001", "商家登录页", "邮箱与密码输入框可正常输入", "login_email_field / login_password_field", "UI显示", "高", "", ""),
        ("M002", "商家登录页", "有效邮箱+密码登录成功 → store_owner 角色跳转 /dashboard", "login_submit_btn", "功能操作", "高", "", ""),
        ("M003", "商家登录页", "brand_admin 角色登录后跳转 /store-selector", "login_submit_btn", "功能操作", "高", "", ""),
        ("M004", "商家登录页", "cashier/trainee 角色登录后跳转 /scan", "login_submit_btn", "功能操作", "高", "", ""),
        ("M005", "商家登录页", "finance 角色登录后跳转 /earnings", "login_submit_btn", "功能操作", "中", "", ""),
        ("M006", "商家登录页", "pending 状态商家登录后跳转 /auth/review", "", "功能操作", "高", "", ""),
        ("M007", "商家登录页", "无商家记录的用户登录后跳转 /auth/register", "", "功能操作", "高", "", ""),
        # ── 注册（6步）──
        ("M008", "商家注册 Step 0", "邮箱格式验证，密码必须≥8字符", "login_email_field / login_password_field", "边界输入", "高", "", ""),
        ("M009", "商家注册 Step 1", "单店/连锁选择正确影响后续步骤显示", "reg_type_single / reg_type_multiple", "UI显示", "高", "", ""),
        ("M010", "商家注册 Step 2", "公司名称必填验证", "register_company_name", "边界输入", "高", "", ""),
        ("M011", "商家注册 Step 2", "连锁模式下品牌名/品牌描述字段出现", "register_brand_name", "UI显示", "高", "", ""),
        ("M012", "商家注册 Step 4", "EIN（税号）字段可输入", "register_ein", "UI显示", "中", "", ""),
        ("M013", "商家注册 Step 5", "地址字段完整填写后可提交", "register_address1 / register_city / register_state / register_zipcode", "边界输入", "高", "", ""),
        ("M014", "商家注册", "重提模式（isResubmit=true）从 Step 2 开始", "", "功能操作", "中", "", ""),
        ("M015", "商家注册", "提交申请后跳转 /auth/review 审核等待页", "register_submit_btn", "功能操作", "高", "", ""),
        # ── 审核状态 ──
        ("M016", "审核状态页", "显示 pending 状态时显示等待审核提示", "", "UI显示", "高", "", ""),
        ("M017", "审核状态页", "显示 rejected 状态时显示拒绝原因并允许重新提交", "", "UI显示", "高", "", ""),
        # ── Dashboard ──
        ("M018", "工作台首页", "4个数据卡片正确显示（今日订单/核销/收入/待处理）", "", "UI显示", "高", "", ""),
        ("M019", "工作台首页", "在线/下线开关切换，乐观更新（失败时回滚）", "", "功能操作", "高", "", ""),
        ("M020", "工作台首页", "8个快捷入口显示正确（品牌管理员额外显示 Brand 入口）", "", "UI显示", "高", "", ""),
        ("M021", "工作台首页", "非品牌管理员不显示 Brand 快捷入口", "isBrandAdmin", "UI显示", "高", "", ""),
        ("M022", "工作台首页", "点击各快捷入口跳转正确路由", "", "功能操作", "高", "", ""),
        ("M023", "工作台首页", "有待确认品牌 Deal 时顶部显示横幅提醒", "", "UI显示", "中", "", ""),
        ("M024", "工作台首页", "下拉刷新更新所有数据", "", "功能操作", "中", "", ""),
        ("M025", "工作台首页", "7日趋势数据列表正确显示", "", "UI显示", "中", "", ""),
        # ── 门店管理 ──
        ("M026", "门店信息页", "正确显示门店名称、描述、电话、地址", "", "UI显示", "高", "", ""),
        ("M027", "门店编辑页", "编辑门店名称/描述/电话/地址，保存成功", "", "功能操作", "高", "", ""),
        ("M028", "营业时间页", "可设置每天开/关状态和时间段", "", "功能操作", "高", "", ""),
        ("M029", "门店照片页", "可上传图片（image_picker），设置主图", "", "功能操作", "中", "", ""),
        ("M030", "门店标签页", "可添加/删除门店特色标签", "", "功能操作", "低", "", ""),
        # ── 员工管理 ──
        ("M031", "员工管理页", "显示当前所有员工及激活状态", "", "UI显示", "中", "", ""),
        ("M032", "员工管理页", "邀请员工弹窗：填写邮箱+角色后发送", "staff_invite_btn / staff_invite_email_field / staff_invite_role_dropdown / staff_invite_submit_btn", "功能操作", "中", "", ""),
        ("M033", "员工管理页", "可激活/禁用/移除员工", "", "功能操作", "中", "", ""),
        # ── 品牌管理 ──
        ("M034", "品牌管理首页", "品牌管理员可见品牌信息头部卡片", "brand_manage_page", "UI显示", "高", "", ""),
        ("M035", "品牌管理首页", "5个功能卡片正确显示（Info/Stores/Admins/Deals/Overview）", "brand_manage_page / brand_manage_back_btn", "UI显示", "高", "", ""),
        ("M036", "品牌信息页", "可编辑品牌名称和描述，保存成功", "", "功能操作", "高", "", ""),
        ("M037", "品牌门店页", "显示旗下所有门店及状态（Active/Pending）", "", "UI显示", "高", "", ""),
        ("M038", "品牌门店页", "通过邮箱邀请新门店加入品牌", "brand_add_store_email_field", "功能操作", "高", "", ""),
        ("M039", "品牌门店页", "点击移除门店显示确认弹窗，确认后移除", "brand_manage_remove_store_*", "功能操作", "高", "", ""),
        ("M040", "品牌管理员页", "显示所有管理员（Owner 角色不可移除）", "", "UI显示", "高", "", ""),
        ("M041", "品牌管理员页", "邀请管理员弹窗：输入邮箱后发送邀请", "brand_admin_email_field / brand_admin_invite_submit_btn", "功能操作", "高", "", ""),
        ("M042", "品牌管理员页", "点击移除非 Owner 管理员，确认后生效", "brand_remove_admin_*", "功能操作", "高", "", ""),
        # ── Deal 管理 ──
        ("M043", "Deal列表页", "4个Tab（All/Active/Inactive/Pending Review）正确筛选", "", "UI显示", "高", "", ""),
        ("M044", "Deal列表页", "点击右下角 + 按钮进入创建 Deal 流程", "", "功能操作", "高", "", ""),
        ("M045", "Deal创建 Step 1", "Deal标题/描述/使用说明字段可填写", "_titleController / _descriptionController / _usageNotesController", "功能操作", "高", "", ""),
        ("M046", "Deal创建 Step 2", "折扣价必须小于原价，否则报错", "_dealPrice", "边界输入", "高", "", ""),
        ("M047", "Deal创建 Step 3", "库存类型：无限制复选框 / 数字限量均可用", "_stockController", "边界输入", "高", "", ""),
        ("M048", "Deal创建 Step 3", "有效期类型：固定日期 / 相对天数切换正确", "_validityDaysController", "UI显示", "高", "", ""),
        ("M049", "Deal创建 Step 4", "品牌管理员可勾选多门店应用范围", "deal_scope_multi_store_btn / deal_store_checkbox_*", "功能操作", "高", "", ""),
        ("M050", "Deal创建 Step 5", "图片上传成功，可设置主图", "", "功能操作", "中", "", ""),
        ("M051", "Deal创建", "完整提交5步后 Deal 状态为 pending review", "", "功能操作", "高", "", ""),
        ("M052", "Deal详情页", "正确显示所有 Deal 字段和当前状态", "", "UI显示", "高", "", ""),
        ("M053", "Deal详情页", "可激活（active）/停用（inactive）Deal", "", "功能操作", "高", "", ""),
        ("M054", "品牌Deal确认页", "门店确认/拒绝品牌 Deal 应用到本店", "", "功能操作", "高", "", ""),
        # ── 扫码核销 ──
        ("M055", "扫码页", "相机权限未授权时显示引导提示", "", "错误处理", "高", "", ""),
        ("M056", "扫码页", "扫描有效 QR 码后跳转确认页", "", "功能操作", "高", "", ""),
        ("M057", "扫码页", "手动输入券码 Tab 可输入并点击 Verify", "", "功能操作", "高", "", ""),
        ("M058", "扫码页", "扫描无效/过期券码显示红色错误 SnackBar", "", "错误处理", "高", "", ""),
        ("M059", "核销确认页", "显示Deal名称、金额、客户信息，点击确认完成核销", "", "功能操作", "高", "", ""),
        ("M060", "核销成功页", "显示成功信息，点击 Done 返回扫码页", "", "UI显示", "高", "", ""),
        ("M061", "核销历史页", "列出历史核销记录（含分页）", "", "UI显示", "中", "", ""),
        # ── 订单 ──
        ("M062", "订单列表页", "4Tab（All/Paid/Redeemed/Refunded）正确筛选", "", "UI显示", "高", "", ""),
        ("M063", "订单列表页", "下拉刷新，上拉加载更多", "", "功能操作", "中", "", ""),
        ("M064", "订单详情页", "正确显示订单信息、核销时间、Deal信息", "", "UI显示", "高", "", ""),
        # ── 财务 ──
        ("M065", "财务主页", "月份选择器正确切换，数据随之更新", "", "功能操作", "高", "", ""),
        ("M066", "财务主页", "4个卡片（总收入/预计结算/已结算/手续费）数据正确", "", "UI显示", "高", "", ""),
        ("M067", "提现页", "提现金额输入并提交", "withdrawal_submit_btn / amountCtrl", "功能操作", "高", "", ""),
        ("M068", "提现页", "自动提现开关可切换", "withdrawal_auto_switch", "功能操作", "中", "", ""),
        ("M069", "收款账户页", "显示 Stripe 连接状态，可发起 onboarding", "", "功能操作", "高", "", ""),
        # ── 设置 ──
        ("M070", "设置页", "Account Security 跳转 /me/account-security（不再404）", "", "功能操作", "高", "", ""),
        ("M071", "设置页", "Staff Accounts 跳转 /me/staff（不再404）", "", "功能操作", "高", "", ""),
        ("M072", "设置页", "Notification Preferences 跳转 /me/notifications（不再404）", "", "功能操作", "高", "", ""),
        ("M073", "设置页", "Help Center 跳转 /me/help（不再404）", "", "功能操作", "高", "", ""),
        ("M074", "设置页", "Sign Out 退出后跳转 /auth/login（不再404）", "", "功能操作", "高", "", "已修复 /login→/auth/login"),
        ("M075", "设置页", "单店商家显示 Upgrade to Chain 入口，连锁商家不显示", "settings_upgrade_chain_btn", "UI显示", "中", "", ""),
        ("M076", "设置页", "升级连锁弹窗：填写品牌名后创建品牌成功", "brandNameCtrl", "功能操作", "中", "", ""),
        ("M077", "设置页", "品牌管理员看到 Brand Management 和 Switch Store 入口", "settings_brand_management_btn", "UI显示", "高", "", ""),
        # ── 菜品管理 ──
        ("M078", "菜品列表页", "正确显示所有菜品及分类", "", "UI显示", "中", "", ""),
        ("M079", "菜品编辑页", "创建/编辑菜品后保存成功", "", "功能操作", "中", "", ""),
        ("M080", "菜品分类管理", "可添加/编辑/删除菜品分类", "", "功能操作", "中", "", ""),
        # ── 门店切换 ──
        ("M081", "门店选择页", "品牌管理员可切换旗下不同门店", "store_selector_item_*", "功能操作", "高", "", ""),
        ("M082", "门店选择页", "切换门店后 Dashboard 数据刷新为新门店数据", "", "功能操作", "高", "", ""),
        # ── 品牌总览 ──
        ("M083", "品牌总览页", "显示所有旗下门店汇总数据", "", "UI显示", "中", "", ""),
        ("M084", "品牌总览页", "可切换查看不同门店数据", "", "功能操作", "中", "", ""),
        # ── 评价 ──
        ("M085", "评价管理页", "显示所有客户评价，含评分和内容", "", "UI显示", "中", "", ""),
        ("M086", "评价管理页", "可对评价进行回复", "", "功能操作", "低", "", ""),
        # ── 通知 ──
        ("M087", "通知页", "显示所有系统通知（未读/已读）", "", "UI显示", "中", "", ""),
        # ── 边界输入 & 错误处理（M088-M103）──
        ("M088", "Deal创建 Step 1", "Deal 标题超过100字符时显示验证错误", "_titleController", "边界输入", "高", "", "max_length=100"),
        ("M089", "Deal创建 Step 2", "折扣价输入0时显示 Price must be > 0 错误", "_dealPrice", "边界输入", "高", "", ""),
        ("M090", "Deal创建 Step 2", "折扣价等于原价时也报错（必须严格小于）", "_dealPrice", "边界输入", "高", "", "边界相等场景"),
        ("M091", "Deal创建 Step 3", "有效天数输入0时显示验证错误", "_validityDaysController", "边界输入", "高", "", ""),
        ("M092", "Deal创建 Step 3", "库存数量输入负数时显示验证错误", "_stockController", "边界输入", "高", "", ""),
        ("M093", "提现页", "提现金额超过当前可提现余额时显示超限错误", "amountCtrl / withdrawal_submit_btn", "边界输入", "高", "", ""),
        ("M094", "提现页", "提现金额为0时提交按钮禁用或显示错误", "amountCtrl", "边界输入", "高", "", ""),
        ("M095", "扫码页", "同一张优惠券被扫码核销两次，第二次显示 Already Redeemed 错误", "", "错误处理", "高", "", "重复核销"),
        ("M096", "扫码页", "扫描已退款券时显示 Coupon Refunded 不可核销提示", "", "错误处理", "高", "", ""),
        ("M097", "扫码页", "扫描 Deal 已过期的优惠券时显示 Deal Expired 提示", "", "错误处理", "高", "", ""),
        ("M098", "员工管理页", "邀请已存在的员工邮箱时显示 Already a staff member 提示", "staff_invite_email_field", "错误处理", "中", "", ""),
        ("M099", "商家注册 Step 4", "EIN 超过20字符时显示格式验证错误", "register_ein", "边界输入", "中", "", ""),
        ("M100", "Deal创建 Step 4", "多门店模式下未勾选任何门店时无法提交到下一步", "deal_scope_multi_store_btn", "边界输入", "高", "", ""),
        ("M101", "设置页", "Close Store 确认后无待退款券时提示 0 vouchers pending refund", "settings_close_store_btn", "边界输入", "中", "", "零边界"),
        ("M102", "商家登录页", "密码超过128字符时正常提示错误，不崩溃", "login_password_field", "边界输入", "中", "", ""),
        ("M103", "设置页", "Upgrade to Chain 弹窗品牌名为空时 Create Brand 按钮不提交", "brandNameCtrl", "边界输入", "高", "", "已有前端验证"),
        # ── Bug #001: 注册邮箱验证缺失 ──
        ("M104", "商家注册 Step 0", "[BUG] 注册后应发送邮箱验证邮件，未验证不允许继续流程", "", "功能操作", "高", "", "安全问题: 当前无邮箱验证步骤"),
        ("M105", "商家注册", "[BUG] 未验证邮箱的账号登录后应提示先完成邮箱验证", "", "功能操作", "高", "", "防止虚假邮箱注册"),
        # ── 需求 #001: 价格异常 Deal 自动检测与停用 ──
        ("M106", "Deal创建 Step 2", "[需求] 原价 < 折扣价时前端阻止提交并显示错误提示", "_dealPrice / _originalPrice", "边界输入", "高", "", "前端校验层"),
        ("M107", "Deal编辑页", "[需求] 编辑时修改价格导致原价 < 折扣价，保存被阻止", "_dealPrice / _originalPrice", "边界输入", "高", "", ""),
        ("M108", "Deal列表页", "[需求] 单店 Deal 因价格异常被后端自动停用后，列表显示 inactive 状态", "", "UI显示", "高", "", "is_active=false"),
        ("M109", "Deal列表页", "[需求] 多店 Deal 部分门店价格异常 → 该门店从 active_store_count 移除，Deal 仍可用", "", "UI显示", "高", "", "仅停用异常门店"),
    ]

    for r_idx, row in enumerate(rows, 2):
        priority = row[5]
        fill = make_fill(HIGH_BG) if priority == "高" else (make_fill(ALT_ROW_BG) if r_idx % 2 == 0 else None)
        write_row(ws, r_idx, row, fill=fill)

    freeze_and_filter(ws)
    return ws

# ─────────────────────────────────────────────────────────
# Sheet 4：后台测试
# ─────────────────────────────────────────────────────────
def build_admin(wb):
    ws = wb.create_sheet("🖥️ 后台管理")
    ws.row_dimensions[1].height = 28

    headers = ["编号", "页面名称", "测试项描述", "代码字段名/Key", "测试类型", "优先级", "测试状态", "备注"]
    widths  = [8, 24, 48, 34, 16, 8, 10, 24]
    write_header(ws, headers, widths)

    rows = [
        # ── 登录 ──
        ("A001", "管理后台登录", "有效邮箱+密码登录，admin 角色跳转 /dashboard", "email / password", "功能操作", "高", "", ""),
        ("A002", "管理后台登录", "非 admin/merchant 账号登录，重定向 /login", "email / password", "功能操作", "高", "", "权限验证"),
        ("A003", "管理后台登录", "无效密码时显示错误提示", "password", "错误处理", "高", "", ""),
        ("A004", "管理后台登录", "侧边栏 Sign out 退出，跳转至 /login", "", "功能操作", "高", "", ""),
        # ── 仪表盘 ──
        ("A005", "仪表盘首页", "Admin 视角：显示用户总数、商家总数、Deal总数、品牌总数", "", "UI显示", "高", "", ""),
        ("A006", "仪表盘首页", "显示待审核商家数量徽章，点击跳转 /merchants", "", "UI显示", "高", "", ""),
        ("A007", "仪表盘首页", "显示待处理退款数量徽章，点击跳转 /orders", "", "UI显示", "高", "", ""),
        # ── 商家审核 ──
        ("A008", "商家列表页", "列出所有商家，显示名称、分类、状态、创建时间", "", "UI显示", "高", "", ""),
        ("A009", "商家列表页", "品牌筛选（All / 各品牌），切换后列表更新", "", "功能操作", "中", "", "URL 参数 ?brand="),
        ("A010", "商家详情页", "显示商家完整信息：名称、联系人、地址、EIN、上传文档", "", "UI显示", "高", "", ""),
        ("A011", "商家详情页", "点击 View / Download 查看并下载上传的营业执照等文档", "", "功能操作", "高", "", ""),
        ("A012", "商家详情页", "点击 Approve 批准商家，状态变为 approved", "approveMerchant", "功能操作", "高", "", ""),
        ("A013", "商家详情页", "点击 Reject → 弹窗输入原因 → 提交，状态变为 rejected，原因保存", "rejectionReason / rejectMerchant", "功能操作", "高", "", ""),
        ("A014", "商家详情页", "点击 Revoke approval → 确认 → 状态回到 pending", "revokeMerchantApproval", "功能操作", "中", "", ""),
        ("A015", "商家详情页", "显示员工列表，可通过 Enable/Disable 切换激活状态", "toggleStaffActive", "功能操作", "中", "", ""),
        # ── 用户管理 ──
        ("A016", "用户管理页", "列出所有用户，显示邮箱、角色、注册时间", "", "UI显示", "中", "", ""),
        ("A017", "用户管理页", "修改用户角色（user/merchant/admin），操作即时生效", "updateUserRole / role", "功能操作", "高", "", ""),
        ("A018", "用户管理页", "当前 admin 自己的角色行标注 admin (you)，不可修改", "", "UI显示", "高", "", ""),
        # ── Deal 管理 ──
        ("A019", "Deal列表页", "列出所有 Deal，显示标题、价格、状态、商家名", "", "UI显示", "高", "", ""),
        ("A020", "Deal列表页", "Sort Order 字段可 inline 编辑并保存", "updateDealSortOrder / value", "功能操作", "中", "", ""),
        ("A021", "Deal详情页", "显示 Deal 完整信息：描述、退款政策、图片、适用门店列表", "", "UI显示", "高", "", ""),
        ("A022", "Deal详情页", "点击 Activate 激活 Deal（is_active=true, deal_status=active）", "setDealActive", "功能操作", "高", "", ""),
        ("A023", "Deal详情页", "点击 Deactivate → 确认弹窗 → 停用 Deal", "setDealActive", "功能操作", "高", "", ""),
        ("A024", "Deal详情页", "多店 Deal 显示适用门店列表及确认状态", "", "UI显示", "中", "", ""),
        # ── 订单管理 ──
        ("A025", "订单列表页", "列出所有订单，显示订单号、Deal、用户、金额、状态", "", "UI显示", "高", "", ""),
        ("A026", "订单列表页", "搜索框输入订单号/邮箱/Deal名，结果实时更新（debounce）", "q / OrderSearchForm", "功能操作", "高", "", ""),
        ("A027", "订单列表页", "Clear 按钮清空搜索，恢复全部列表", "", "功能操作", "中", "", ""),
        ("A028", "订单列表页", "refund_requested 状态订单高亮显示（橙色背景）", "", "UI显示", "高", "", ""),
        ("A029", "订单详情页", "显示完整订单信息、核销门店（如果与下单门店不同）、退款信息", "", "UI显示", "高", "", ""),
        ("A030", "订单详情页", "退款请求：点击 Approve 批准退款，status → refunded", "approveRefund", "功能操作", "高", "", ""),
        ("A031", "订单详情页", "退款请求：点击 Reject 拒绝退款，status → used", "rejectRefund", "功能操作", "高", "", ""),
        # ── 品牌管理 ──
        ("A032", "品牌列表页", "列出所有品牌，显示名称、Logo、门店数、管理员数", "", "UI显示", "中", "", ""),
        ("A033", "品牌详情页", "Member Stores 表格：显示旗下门店，可移除门店", "removeStoreFromBrand", "功能操作", "高", "", ""),
        ("A034", "品牌详情页", "Add Store 下拉选择门店并点击 Add，门店加入品牌", "addStoreToBrand / selectedId", "功能操作", "高", "", ""),
        ("A035", "品牌详情页", "Brand Admins 表格：显示管理员，可移除", "removeBrandAdmin", "功能操作", "高", "", ""),
        ("A036", "品牌详情页", "Add Brand Admin：填写邮箱+角色后发送邀请", "addBrandAdmin / email / role", "功能操作", "高", "", ""),
        ("A037", "品牌详情页", "Invitations 表格：显示待接受的邀请记录", "", "UI显示", "中", "", ""),
        ("A038", "品牌详情页", "All Staff 表格：显示品牌所有门店的员工", "", "UI显示", "低", "", ""),
        # ── 财务 ──
        ("A039", "财务页", "显示提现记录表格（商家名、品牌、金额、状态、日期）", "", "UI显示", "中", "", ""),
        ("A040", "财务页", "品牌筛选（All / 各品牌），显示对应品牌提现汇总", "", "功能操作", "中", "", ""),
        # ── Closures ──
        ("A041", "关闭记录页", "显示 Closed Stores 表格（含总退款金额）", "", "UI显示", "中", "", ""),
        ("A042", "关闭记录页", "显示 Rejected Merchants 和 Brand Disassociations 记录", "", "UI显示", "低", "", ""),
        # ── 边界输入 & 错误处理（A043-A057）──
        ("A043", "商家详情页", "Reject 弹窗中拒绝原因字段为空时，提交被阻止并显示错误", "rejectionReason / rejectMerchant", "边界输入", "高", "", ""),
        ("A044", "商家详情页", "先 Reject 后再次点击 Approve，商家状态恢复 approved", "approveMerchant / rejectMerchant", "功能操作", "高", "", "状态流转"),
        ("A045", "Deal列表页", "Sort Order 输入非数字字符时显示格式错误提示", "updateDealSortOrder / value", "边界输入", "中", "", ""),
        ("A046", "Deal列表页", "Sort Order 输入负数时显示验证错误", "updateDealSortOrder / value", "边界输入", "中", "", ""),
        ("A047", "订单列表页", "搜索框输入 SQL 注入字符串（' OR 1=1--）不崩溃，正常返回空结果", "q / OrderSearchForm", "错误处理", "高", "", "安全性"),
        ("A048", "订单详情页", "对已经是 refunded 状态的订单点击 Approve Refund 时显示冲突错误", "approveRefund", "错误处理", "高", "", "重复退款"),
        ("A049", "品牌详情页", "Add Store 时选择已在该品牌内的门店，显示 Already in brand 提示", "addStoreToBrand / selectedId", "错误处理", "中", "", ""),
        ("A050", "品牌详情页", "移除品牌内最后一个门店时显示警告确认框", "removeStoreFromBrand", "边界输入", "中", "", "零门店边界"),
        ("A051", "品牌详情页", "Add Brand Admin 填写不存在的邮箱时显示 User not found 错误", "addBrandAdmin / email", "错误处理", "中", "", ""),
        ("A052", "用户管理页", "将当前登录的 admin 角色改为 user 时显示自降权警告", "updateUserRole / role", "错误处理", "高", "", "权限自锁"),
        ("A053", "用户管理页", "将当前登录的 admin 角色改为 merchant 时显示确认提示", "updateUserRole / role", "错误处理", "高", "", ""),
        ("A054", "商家审核", "同时快速双击 Approve 按钮，不出现重复审批或状态异常", "approveMerchant", "错误处理", "中", "", "并发防重"),
        ("A055", "财务页", "品牌筛选传入不存在的品牌 ID 时页面不崩溃，显示空状态", "", "错误处理", "中", "", ""),
        ("A056", "管理后台", "会话过期（JWT 失效）后操作页面，重定向至 /login 而非白屏", "", "错误处理", "高", "", ""),
        ("A057", "Deal详情页", "对 deal_status=deleted 的软删除 Deal 点击 Activate 时显示操作不可用", "setDealActive", "错误处理", "中", "", ""),
        # ── 需求 #001: 价格异常 Deal 自动检测 ──
        ("A058", "Deal详情页", "[需求] 后台激活 Deal 时检测原价<折扣价，显示价格异常警告并阻止激活", "setDealActive", "边界输入", "高", "", "审核激活时触发"),
        ("A059", "Deal列表页", "[需求] 价格异常的 Deal 显示警告标记（如红色感叹号）", "", "UI显示", "高", "", "方便运营快速识别"),
        ("A060", "Deal详情页", "[需求] 多店 Deal 价格异常时显示哪些门店有异常", "", "UI显示", "中", "", "精确到门店"),
    ]

    for r_idx, row in enumerate(rows, 2):
        priority = row[5]
        fill = make_fill(HIGH_BG) if priority == "高" else (make_fill(ALT_ROW_BG) if r_idx % 2 == 0 else None)
        write_row(ws, r_idx, row, fill=fill)

    freeze_and_filter(ws)
    return ws

# ─────────────────────────────────────────────────────────
# Sheet 5：三端联动测试
# ─────────────────────────────────────────────────────────
def build_cross(wb):
    ws = wb.create_sheet("🔗 三端联动")
    ws.row_dimensions[1].height = 28

    # 10列固定结构：编号|场景名称|步骤1|步骤2|步骤3|预期最终结果|涉及字段/Key|优先级|测试状态|备注
    headers = [
        "编号",
        "场景名称",
        "步骤1（端+操作）",
        "步骤2（端+操作）",
        "步骤3（端+操作）",
        "预期最终结果",
        "涉及字段/Key",
        "优先级",
        "测试状态",
        "备注",
    ]
    widths = [8, 30, 36, 36, 36, 40, 36, 8, 10, 20]
    write_header(ws, headers, widths)

    # 每行严格10个元素，顺序对应上方10列
    # (编号, 场景名称, 步骤1, 步骤2, 步骤3, 预期结果, 涉及字段, 优先级, 状态, 备注)
    rows = [
        (
            "X001",
            "商家注册 → 后台审核 → 商家可登录",
            "【商家端】商家填写注册信息（6步）并提交申请",
            "【后台】Merchants 列表显示待审核商家，Admin 点击 Approve 批准",
            "",
            "商家端：状态变为 approved；商家重新登录后跳转 /dashboard 而非 /auth/review",
            "merchants.status; approveMerchant(); MerchantStatusCache",
            "高", "", "核心流程",
        ),
        (
            "X002",
            "商家注册被拒绝 → 商家看到原因 → 重新提交",
            "【商家端】商家提交注册申请",
            "【后台】Admin 输入拒绝原因后点击 Reject",
            "【商家端】商家登录，审核状态页显示拒绝原因，点击重提",
            "商家端：MerchantRegisterPage 从 Step 2 开始（isResubmit=true），重提后状态回到 pending",
            "rejectionReason; merchants.status=rejected; isResubmit",
            "高", "", "",
        ),
        (
            "X003",
            "商家发布 Deal → 后台审核激活 → 客户端可见",
            "【商家端】商家完成 5 步 Deal 创建，提交后状态为 pending_review",
            "【后台】Deals 列表显示该 Deal，Admin 点击 Activate",
            "",
            "客户端首页：该 Deal 出现在列表中（is_active=true, deal_status=active）",
            "deals.is_active; deals.deal_status; setDealActive(); dealsProvider",
            "高", "", "",
        ),
        (
            "X004",
            "商家停用 Deal → 客户端立即不可见",
            "【商家端】商家在 Deal 详情页点击 Deactivate",
            "",
            "",
            "客户端首页/搜索：该 Deal 不再显示；Deal 详情页跳转404或提示已下架",
            "deals.is_active=false; dealsProvider; HomeScreen",
            "高", "", "",
        ),
        (
            "X005",
            "客户下单 → 商家端接收订单 → 后台可查",
            "【客户端】用户选择 Deal → 支付成功 → 跳转 /order-success",
            "【商家端】Dashboard 今日订单+1；订单列表出现新订单",
            "【后台】Orders 列表显示该订单（状态 unused）",
            "三端均可看到该订单；订单状态 unused",
            "orders表; dashboardProvider; OrdersListPage; admin /orders",
            "高", "", "核心流程",
        ),
        (
            "X006",
            "客户核销优惠券 → 商家扫码确认 → 订单状态变更",
            "【客户端】用户在 /coupon/:id 显示 QR 码",
            "【商家端】收银员在 ScanPage 扫描 QR → 确认核销",
            "",
            "客户端：订单状态变为 used；商家端：核销成功页显示；后台：orders.status=used",
            "merchant-scan Edge Function; coupons.redeemed_at; orders.status=used",
            "高", "", "",
        ),
        (
            "X007",
            "客户申请退款 → 后台审批 → 客户端状态更新",
            "【客户端】用户在 /refund/:orderId 提交退款申请",
            "【后台】Orders 列表出现 refund_requested 高亮订单，Admin 点击 Approve",
            "",
            "客户端：优惠券页状态变为 Refunded；后台：orders.status=refunded",
            "refund_reason; approveRefund; orders.status=refund_requested→refunded",
            "高", "", "",
        ),
        (
            "X008",
            "后台拒绝退款 → 客户端状态不变",
            "【客户端】用户提交退款申请",
            "【后台】Admin 点击 Reject 拒绝退款",
            "",
            "客户端：订单状态回到 used（不是 refunded）",
            "rejectRefund; orders.status=used",
            "高", "", "",
        ),
        (
            "X009",
            "后台封禁商家（Revoke） → 商家端无法进入主功能 → 客户端活动下架",
            "【后台】Admin 对 approved 商家点击 Revoke approval",
            "【商家端】商家尝试访问 /dashboard，路由守卫检测 status→pending，跳转 /auth/review",
            "【客户端】该商家所有活跃 Deal 自动下架（is_active=false）",
            "商家端：重定向到审核等待页；客户端：该商家 Deal 消失",
            "revokeMerchantApproval; MerchantStatusCache; deals.is_active",
            "高", "", "需确认客户端下架是否自动触发",
        ),
        (
            "X010",
            "商家修改 Deal 价格 → 客户端实时显示新价格",
            "【商家端】商家在 Deal 编辑页修改折扣价并保存",
            "",
            "",
            "客户端：Deal 列表和详情页显示新价格",
            "deals.discount_price; merchant-deals PATCH; dealsProvider",
            "高", "", "",
        ),
        (
            "X011",
            "单店升级为连锁（Upgrade to Chain）→ 后台可见品牌 → 品牌管理功能解锁",
            "【商家端】设置页点击 Upgrade to Chain，填写品牌名后创建",
            "【后台】Brands 列表出现新品牌",
            "【商家端】Dashboard Quick Actions 出现 Brand 入口；Settings 出现 Brand Management",
            "brands 表出现新记录；商家端 isBrandAdmin=true；storeProvider 刷新后 isChainStore=true",
            "createBrand; brands; isBrandAdmin; ShortcutGrid; settings_brand_management_btn",
            "高", "", "",
        ),
        (
            "X012",
            "品牌管理员从后台/商家端邀请门店 → 被邀请门店确认 → 品牌关联生效",
            "【后台 或 商家端】品牌管理员邀请门店（通过邮箱）",
            "【被邀请门店商家端】收到通知，点击接受邀请",
            "【后台】Brands 详情页 Member Stores 表格新增该门店",
            "品牌门店关联成功；被邀请门店 isChainStore=true",
            "brand_invitations; addStoreToBrand; selectedId; isChainStore",
            "中", "", "需确认邀请通知机制",
        ),
        (
            "X013",
            "商家发布品牌多店 Deal → 各门店收到确认请求 → 客户可在多门店使用",
            "【商家端-品牌管理员】创建 Deal 时选择多门店，提交并经后台激活",
            "【商家端-子门店】Dashboard 显示待确认 Deal 横幅，进入确认页后点击确认",
            "【客户端】Deal 详情页显示多门店门店数",
            "多门店 Deal 在所有已确认门店可使用",
            "deal_applicable_stores; deal_scope_multi_store_btn; StoreDealConfirmPage; active_store_count",
            "高", "", "核心多店功能",
        ),
        (
            "X014",
            "子门店拒绝品牌 Deal → 该门店不出现在适用范围",
            "【商家端-品牌管理员】创建并激活多店 Deal",
            "【商家端-子门店】进入确认页，点击拒绝",
            "【客户端】Deal 详情页门店列表不包含拒绝的门店",
            "拒绝门店不在适用范围；active_store_count 减少",
            "deal_applicable_stores.confirmed=false; active_store_count",
            "高", "", "",
        ),
        (
            "X015",
            "客户在子门店核销品牌 Deal → 商家端（该子门店）记录",
            "【客户端】用户出示优惠券 QR",
            "【商家端-子门店】扫码核销",
            "",
            "商家端核销成功；后台订单详情显示核销门店（与下单门店可能不同）",
            "coupons.redeemed_at_merchant_id; orders scan log",
            "高", "", "",
        ),
        (
            "X016",
            "商家关闭店铺 → 活跃 Deal 下架 → 未使用券退款",
            "【商家端】设置页点击 Close Store → 确认",
            "",
            "【后台】Closures 页出现该关闭记录；退款统计更新",
            "商家状态 Closed；所有 active Deal 变 inactive；未使用券触发退款",
            "closeStore; settings_close_store_btn; closures; auto-refund-expired Edge Function",
            "高", "", "需确认退款是否自动触发",
        ),
        (
            "X017",
            "后台修改用户角色为 merchant → 用户使用商家端登录",
            "【后台】Users 页将目标用户角色从 user 改为 merchant",
            "【商家端】该用户登录，进入注册流程",
            "",
            "商家端：路由守卫检测 status=none → 跳转 /auth/register",
            "users.role; updateUserRole; role; MerchantStatusCache",
            "中", "", "",
        ),
        (
            "X018",
            "商家端员工账号登录 → 权限限制验证",
            "【商家端】store_owner 在员工管理页邀请 cashier 角色员工",
            "【商家端-员工账号】cashier 登录，验证跳转至 /scan",
            "",
            "cashier 员工只能访问 /scan；无法访问 /deals、/earnings 等",
            "merchant_staff; role=cashier; 路由守卫 roleType; staff_invite_role_dropdown",
            "高", "", "",
        ),
        # ── 需求 #001: 价格异常 Deal 自动检测（三端联动）──
        (
            "X019",
            "[需求] 单店 Deal 价格异常 → 后端自动停用 → 客户端不可见",
            "【商家端】商家创建 Deal（原价=50，折扣价=80，原价<折扣价）",
            "【后端】Edge Function 或定时巡检检测到价格异常，自动设置 is_active=false",
            "【客户端】该 Deal 从首页/搜索消失",
            "Deal 被自动停用；商家端显示 inactive；客户端不可见",
            "deals.original_price; deals.deal_price; deals.is_active; auto-price-check",
            "高", "", "单店场景",
        ),
        (
            "X020",
            "[需求] 多店 Deal 部分门店价格异常 → 仅停用异常门店 → 其他门店正常",
            "【商家端-品牌管理员】创建多门店 Deal（3家门店，其中1家原价<折扣价）",
            "【后端】检测到1家门店价格异常，将该门店从 active_store_count 移除",
            "【客户端】Deal 仍可见，但门店列表不包含异常门店，active_store_count=2",
            "异常门店被移除；Deal 本身仍活跃；active_store_count 正确",
            "deal_applicable_stores; active_store_count; price_anomaly_check",
            "高", "", "多店场景",
        ),
    ]

    client_fill   = make_fill(CLIENT_BG)
    merchant_fill = make_fill(MERCHANT_BG)
    admin_fill    = make_fill(ADMIN_BG)
    high_fill     = make_fill(HIGH_BG)
    alt_fill      = make_fill(ALT_ROW_BG)

    for r_idx, row in enumerate(rows, 2):
        # 严格校验：每行必须10个元素
        assert len(row) == 10, f"Row {row[0]} has {len(row)} elements, expected 10"
        priority = row[7]   # 第8列（0-based index 7）= 优先级
        base_fill = high_fill if priority == "高" else (alt_fill if r_idx % 2 == 0 else None)
        write_row(ws, r_idx, row, fill=base_fill)
        # 对步骤列（col 3/4/5）根据内容单独着色
        for col_idx in [3, 4, 5]:
            cell = ws.cell(row=r_idx, column=col_idx)
            step_val = cell.value or ""
            if "客户端" in step_val:
                cell.fill = client_fill
            elif "商家端" in step_val:
                cell.fill = merchant_fill
            elif "后台" in step_val:
                cell.fill = admin_fill
            # 空步骤保留基础行颜色

    freeze_and_filter(ws)
    return ws

# ─────────────────────────────────────────────────────────
# 主流程
# ─────────────────────────────────────────────────────────
def main():
    wb = Workbook()
    # 删除默认 Sheet
    del wb[wb.sheetnames[0]]

    build_overview(wb)
    build_client(wb)
    build_merchant(wb)
    build_admin(wb)
    build_cross(wb)

    output_path = "/Users/howardshansmac/github/coupon app/coupon-app/test_plan.xlsx"
    wb.save(output_path)

    # 统计各 Sheet 行数
    counts = {ws.title: ws.max_row - 1 for ws in wb.worksheets}
    print(f"✅ 已生成: {output_path}")
    for name, cnt in counts.items():
        print(f"   {name}: {cnt} 条测试项")

if __name__ == "__main__":
    main()
